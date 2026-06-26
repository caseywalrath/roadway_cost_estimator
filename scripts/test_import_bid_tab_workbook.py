from __future__ import annotations

import tempfile
import unittest
import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parent))

from import_bid_tab_workbook import (
    DEFAULT_PROJECT_ID,
    DEFAULT_ROW_PREFIX,
    DEFAULT_SOURCE_ID,
    DEFAULT_WORKBOOK,
    agency_item_rows,
    normalize_unit,
    parse_workbook,
    resolve_short_item_codes,
    validate_import,
)


class BidTabWorkbookImportTests(unittest.TestCase):
    def test_unit_normalization_preserves_raw_unknowns(self) -> None:
        self.assertEqual("EACH", normalize_unit("EA"))
        self.assertEqual("L S", normalize_unit("Lump Sum"))
        self.assertEqual("LB", normalize_unit("POUNDS"))
        self.assertEqual("SY", normalize_unit("Sq Yd"))
        self.assertEqual("HOUR", normalize_unit("HR"))
        self.assertEqual("DAY", normalize_unit("DY"))
        self.assertEqual("ACRE", normalize_unit("AC"))
        self.assertEqual("F A", normalize_unit("F/A"))

    def test_parser_extracts_saq_bid_tab_shape(self) -> None:
        workbook_path = self.write_sample_workbook()

        parsed = parse_workbook(
            workbook_path,
            "sample_source",
            "sample_project",
            "sample_prefix",
            "2021-12-16",
        )

        self.assertEqual("Watson Avenue Roundabout", parsed.project_name)
        self.assertEqual("320150", parsed.project_number)
        self.assertEqual(3, len(parsed.item_rows))
        self.assertEqual(2, len(parsed.bidder_bids))
        self.assertEqual(6, len(parsed.observations))
        self.assertEqual(6, len(parsed.bidder_items))

        totals = {row["bidder_name"]: row["bid_total"] for row in parsed.bidder_bids}
        self.assertEqual("755.00", totals["BIDDER A"])
        self.assertEqual("925.00", totals["BIDDER B"])
        self.assertEqual("true", parsed.bidder_bids[0]["apparent_low"])
        self.assertEqual("false", parsed.bidder_bids[1]["apparent_low"])

        repeated_code_rows = [row for row in parsed.item_rows if row["item_code"] == "619-00000"]
        self.assertEqual(2, len(repeated_code_rows))
        self.assertEqual("sample_prefix_row_0019", parsed.item_rows[0]["bid_tab_item_id"])
        self.assertEqual("sample_prefix_row_0019", parsed.bidder_items[0]["bid_tab_item_id"])
        observation_ids = {row["observation_id"] for row in parsed.observations}
        self.assertIn("sample_prefix_row_0020_average", observation_ids)
        self.assertIn("sample_prefix_row_0021_average", observation_ids)

    def test_parser_extracts_bid_form_shape(self) -> None:
        workbook_path = self.write_bid_form_sample_workbook()

        parsed = parse_workbook(
            workbook_path,
            "sample_source",
            "sample_project",
            "sample_bid_form",
            "2021-09-24",
        )

        self.assertEqual("Arapahoe Rd Bridge at Big Dry Creek Project", parsed.project_name)
        self.assertEqual("IFB 21-09-02", parsed.project_number)
        self.assertEqual(3, len(parsed.item_rows))
        self.assertEqual(5, len(parsed.bidder_bids))
        self.assertEqual(6, len(parsed.observations))
        self.assertEqual(15, len(parsed.bidder_items))

        totals = {row["bidder_name"]: row["bid_total"] for row in parsed.bidder_bids}
        self.assertEqual("100.00", totals["HAMON"])
        self.assertEqual("120.00", totals["JALISCO"])
        self.assertEqual("140.00", totals["ZAK DIRT"])
        self.assertEqual("160.00", totals["FLATIRON"])
        self.assertEqual("180.00", totals["ACC"])
        self.assertEqual("HAMON", parsed.bidder_bids[0]["bidder_name"])
        self.assertEqual("true", parsed.bidder_bids[0]["apparent_low"])

        first_average = next(row for row in parsed.item_rows if row["workbook_row"] == "2")
        self.assertEqual("14.00", first_average["average_bid_unit_price"])
        self.assertEqual("sample_bid_form_row_0002", first_average["bid_tab_item_id"])
        self.assertEqual("HOUR", parsed.item_rows[1]["unit_normalized"])
        self.assertEqual("F A", parsed.item_rows[2]["unit_normalized"])
        self.assertIn("Sheet1 engineer estimate 90.00 differs from itemized engineer estimate total 100.00", parsed.warnings)

    def test_validation_blocks_missing_item_codes_without_flag(self) -> None:
        parsed = parse_workbook(
            self.write_sample_workbook(),
            "sample_source",
            "sample_project",
            "sample_prefix",
            "2021-12-16",
        )

        errors, warnings = validate_import(parsed, {"201-00000"}, allow_missing_item_codes=False)

        self.assertIn("Missing agency item codes: 619-00000", errors)
        self.assertEqual([], warnings)

    def test_validation_warns_for_missing_item_codes_with_flag(self) -> None:
        parsed = parse_workbook(
            self.write_sample_workbook(),
            "sample_source",
            "sample_project",
            "sample_prefix",
            "2021-12-16",
        )

        errors, warnings = validate_import(parsed, {"201-00000"}, allow_missing_item_codes=True)

        self.assertEqual([], errors)
        self.assertIn("Missing agency item codes: 619-00000", warnings)

    @unittest.skipUnless(DEFAULT_WORKBOOK.exists(), "Watson workbook is not present on this machine.")
    def test_real_watson_workbook_parse_contract(self) -> None:
        parsed = parse_workbook(
            DEFAULT_WORKBOOK,
            DEFAULT_SOURCE_ID,
            DEFAULT_PROJECT_ID,
            DEFAULT_ROW_PREFIX,
            "2021-12-16",
        )

        self.assertEqual(174, len(parsed.item_rows))
        self.assertEqual(2, len(parsed.bidder_bids))
        totals = {row["bidder_name"]: row["bid_total"] for row in parsed.bidder_bids}
        self.assertEqual("5060815.50", totals["COLUMBINE HILLS CONSTRUCTION"])
        self.assertEqual("5305622.20", totals["KRAEMER NORTH AMERICA"])
        self.assertEqual("COLUMBINE HILLS CONSTRUCTION", parsed.bidder_bids[0]["bidder_name"])
        self.assertEqual("true", parsed.bidder_bids[0]["apparent_low"])

    @unittest.skipUnless(
        Path(r"C:\Users\Casey.Walrath\Downloads\2021 09 21 Bids Arapahoe Road over Big Dry Creek_BID FORM.xlsx").exists(),
        "Arapahoe workbook is not present on this machine.",
    )
    def test_real_arapahoe_workbook_parse_contract(self) -> None:
        parsed = parse_workbook(
            Path(r"C:\Users\Casey.Walrath\Downloads\2021 09 21 Bids Arapahoe Road over Big Dry Creek_BID FORM.xlsx"),
            "fhu_bid_tab_arapahoe_big_dry_creek_2021_09_24",
            "fhu_bid_tab_arapahoe_big_dry_creek_2021_09_24_ifb_21_09_02",
            "fhu_arapahoe_big_dry_creek_20210924",
            "2021-09-24",
        )

        self.assertEqual("Arapahoe Rd Bridge at Big Dry Creek Project", parsed.project_name)
        self.assertEqual("IFB 21-09-02", parsed.project_number)
        self.assertEqual(172, len(parsed.item_rows))
        self.assertEqual(5, len(parsed.bidder_bids))
        totals = {row["bidder_name"]: row["bid_total"] for row in parsed.bidder_bids}
        self.assertEqual("8626000.00", totals["HAMON"])
        self.assertEqual("9286393.43", totals["JALISCO"])
        self.assertEqual("9374747.85", totals["ZAK DIRT"])
        self.assertEqual("9440701.94", totals["FLATIRON"])
        self.assertEqual("10549863.00", totals["ACC"])
        self.assertEqual("HAMON", parsed.bidder_bids[0]["bidder_name"])
        self.assertEqual("true", parsed.bidder_bids[0]["apparent_low"])
        self.assertEqual(["208", "216", "506", "601", "619", "642", "700"], sorted({row["item_code"] for row in parsed.item_rows if "-" not in row["item_code"]}))
        self.assertIn(
            "Sheet1 engineer estimate 7000000.00 differs from itemized engineer estimate total 7085062.75",
            parsed.warnings,
        )

    @unittest.skipUnless(
        Path(r"C:\Users\Casey.Walrath\Downloads\2021 02 05 Ralston Rd Yukon to Garrison Bid Tab.xlsx").exists(),
        "Ralston workbook is not present on this machine.",
    )
    def test_real_ralston_workbook_parse_contract(self) -> None:
        parsed = parse_workbook(
            Path(r"C:\Users\Casey.Walrath\Downloads\2021 02 05 Ralston Rd Yukon to Garrison Bid Tab.xlsx"),
            "fhu_bid_tab_ralston_yukon_garrison_2021_02_05",
            "fhu_bid_tab_ralston_yukon_garrison_2021_02_05_18_st_40",
            "fhu_ralston_yukon_garrison_20210205",
            "2021-02-05",
        )

        self.assertEqual("Ralston Road - Yukon Street to Garrison Street", parsed.project_name)
        self.assertEqual("18-ST-40", parsed.project_number)
        self.assertEqual(235, len(parsed.item_rows))
        self.assertEqual(6, len(parsed.bidder_bids))
        self.assertEqual(1410, len(parsed.bidder_items))
        self.assertEqual([], parsed.observations)

        totals = {row["bidder_name"]: row["bid_total"] for row in parsed.bidder_bids}
        self.assertEqual("12360000.00", totals["Hamon Infrastructure"])
        self.assertEqual("13360995.50", totals["Edge Contracting"])
        self.assertEqual("14079260.56", totals["Brannan Sand & Gravel Co"])
        self.assertEqual("14115478.07", totals["Concrete Works of Colorado, Inc."])
        self.assertEqual("14789252.31", totals["Flatiron Constructors, Inc."])
        self.assertEqual("15041491.00", totals["American Civil Constructors, Inc"])
        self.assertEqual("Hamon Infrastructure", parsed.bidder_bids[0]["bidder_name"])
        self.assertEqual("true", parsed.bidder_bids[0]["apparent_low"])

        blank_spec_rows = [row for row in parsed.item_rows if not row["source_spec_raw"]]
        self.assertEqual(9, len(blank_spec_rows))
        units = {row["unit_raw"]: row["unit_normalized"] for row in parsed.item_rows}
        self.assertEqual("DAY", units["DY"])
        self.assertEqual("HOUR", units["HR"])
        self.assertEqual("F A", units["FA"])
        self.assertEqual("ACRE", units["AC"])
        self.assertTrue(all(row["match_status"] != "matched" for row in parsed.item_rows))

    @unittest.skipUnless(
        Path(r"C:\Users\Casey.Walrath\Downloads\Ralston_Rd_CDOT_reconciliation.xlsx").exists(),
        "Ralston reconciliation workbook is not present on this machine.",
    )
    def test_real_ralston_reconciliation_workbook_parse_contract(self) -> None:
        parsed = parse_workbook(
            Path(r"C:\Users\Casey.Walrath\Downloads\Ralston_Rd_CDOT_reconciliation.xlsx"),
            "fhu_bid_tab_ralston_yukon_garrison_2021_02_05",
            "fhu_bid_tab_ralston_yukon_garrison_2021_02_05_18_st_40",
            "fhu_ralston_yukon_garrison_20210205",
            "2021-02-05",
        )

        matched_rows = [row for row in parsed.item_rows if row["match_status"] == "matched"]
        unmatched_rows = [row for row in parsed.item_rows if row["match_status"] == "unmatched"]
        agency_codes = {
            row["item_code"].strip().upper()
            for row in agency_item_rows(Path("public/data/agency_items.csv"))
            if row.get("item_code")
        }

        self.assertEqual(235, len(parsed.item_rows))
        self.assertEqual(210, len(matched_rows))
        self.assertEqual(25, len(unmatched_rows))
        self.assertEqual(420, len(parsed.observations))
        self.assertTrue(all(row["item_code"] in agency_codes for row in matched_rows))
        self.assertEqual("COA 3.2", parsed.item_rows[0]["source_item_code"])
        self.assertEqual("201-00000", parsed.item_rows[0]["matched_agency_item_code"])
        self.assertEqual("matched", parsed.item_rows[0]["match_status"])

        first_observation = next(row for row in parsed.observations if row["observation_id"].endswith("_average"))
        self.assertEqual("201-00000", first_observation["agency_item_code"])
        self.assertEqual("Clearing and Grubbing", first_observation["description_raw"])
        self.assertEqual("L S", first_observation["unit_normalized"])

        self.assertTrue(all(row["matched_agency_item_code"] == "" for row in unmatched_rows))
        self.assertTrue(all(row["agency_item_code"] == "" for row in parsed.bidder_items if row["bid_tab_item_id"] in {item["bid_tab_item_id"] for item in unmatched_rows}))

    @unittest.skipUnless(
        Path(r"C:\Users\Casey.Walrath\Downloads\2022 03 10 Contractor Bids Kipling at Bowles(1).xlsx").exists(),
        "Kipling/Bowles workbook is not present on this machine.",
    )
    def test_real_kipling_bowles_workbook_parse_contract(self) -> None:
        parsed = parse_workbook(
            Path(r"C:\Users\Casey.Walrath\Downloads\2022 03 10 Contractor Bids Kipling at Bowles(1).xlsx"),
            "fhu_bid_tab_kipling_bowles_2022_03_10",
            "fhu_bid_tab_kipling_bowles_2022_03_10_5_69_15_4005",
            "fhu_kipling_bowles_20220310",
            "2022-03-10",
        )

        agency_codes = {
            row["item_code"].strip().upper()
            for row in agency_item_rows(Path("public/data/agency_items.csv"))
            if row.get("item_code")
        }
        totals = {row["bidder_name"]: row["bid_total"] for row in parsed.bidder_bids}

        self.assertEqual("S Kipling Pkwy Improvement Project at W Bowles Ave", parsed.project_name)
        self.assertEqual("5-69-15-4005", parsed.project_number)
        self.assertEqual(108, len(parsed.item_rows))
        self.assertEqual(4, len(parsed.bidder_bids))
        self.assertEqual(432, len(parsed.bidder_items))
        self.assertEqual(216, len(parsed.observations))
        self.assertEqual([], parsed.warnings)
        self.assertEqual("SEMA", parsed.bidder_bids[0]["bidder_name"])
        self.assertEqual("true", parsed.bidder_bids[0]["apparent_low"])
        self.assertEqual("2387458.90", totals["SEMA"])
        self.assertEqual("2487594.00", totals["American West"])
        self.assertEqual("2540424.10", totals["CEI"])
        self.assertEqual("2629502.45", totals["EDGE"])
        self.assertEqual("60357.50", parsed.item_rows[0]["average_bid_unit_price"])
        self.assertEqual("700-70380", parsed.item_rows[-1]["item_code"])
        self.assertTrue(all(row["item_code"] in agency_codes for row in parsed.item_rows))
        self.assertFalse(any(int(row["workbook_row"]) >= 120 for row in parsed.item_rows))

    def test_short_item_code_resolver_only_applies_exact_compatible_matches(self) -> None:
        parsed = parse_workbook(
            self.write_bid_form_sample_workbook(),
            "sample_source",
            "sample_project",
            "sample_bid_form",
            "2021-09-24",
        )

        warnings = resolve_short_item_codes(
            parsed,
            [
                {
                    "item_code": "700-70310",
                    "official_description": "F/A Landscaping",
                    "official_abbreviated_description": "F/A Landscaping",
                    "official_unit": "F A",
                },
                {
                    "item_code": "700-70072",
                    "official_description": "F/A Landscaping Restoration",
                    "official_abbreviated_description": "F/A Landscaping Restoration",
                    "official_unit": "F A",
                },
            ],
        )

        self.assertEqual("700", parsed.item_rows[2]["item_code"])
        self.assertIn(
            "Preserved nonstandard item code 700 for F/A LANDSCAPE RESTORATION (F A); no exact compatible CDOT item match.",
            warnings,
        )

    def write_sample_workbook(self) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SAQ"

        sheet["C8"] = "SUMMARY OF QUANTITIES - ENGINEER'S ESTIMATE AND BID TABULATION"
        sheet["C10"] = "PROJECT NAME: WATSON AVENUE ROUNDABOUT"
        sheet["C11"] = "PROJECT NO.: 320150"
        sheet["G16"] = "ENGINEER'S ESTIMATE"
        sheet["I16"] = "BIDDER A"
        sheet["K16"] = "BIDDER B"
        sheet["M16"] = "AVERAGE BID"
        headers = ["ITEM CODE", "DESCRIPTION", "QUANTITY", "UNIT", "UNIT COST", "TOTAL COST", "UNIT COST", "TOTAL COST", "UNIT COST", "TOTAL COST", "UNIT COST", "TOTAL COST"]
        for index, value in enumerate(headers, start=3):
            sheet.cell(17, index).value = value

        rows = [
            ["201-00000", "CLEARING AND GRUBBING", 1, "LS", 100, 100, 120, 120, 140, 140, 130, 130],
            ["619-00000", "16 INCH X 6 INCH TEE", 5, "EA", 50, 250, 75, 375, 85, 425, 80, 400],
            ["619-00000", "THRUST BLOCK", 10, "POUNDS", 2, 20, 26, 260, 36, 360, 31, 310],
        ]
        for row_index, row in enumerate(rows, start=19):
            for column_index, value in enumerate(row, start=3):
                sheet.cell(row_index, column_index).value = value

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        workbook.save(tmp.name)
        return Path(tmp.name)

    def write_bid_form_sample_workbook(self) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Bid Form"
        headers = [
            "CDOT No.",
            "ITEM NO.",
            "ITEM DESCRIPTION",
            "UNIT",
            "QUANTITY",
            "ESTIMATED UNIT PRICE",
            "TOTAL COST BASE BID",
            "HAMON UNIT PRICE",
            "TOTAL COST BASE BID",
            "JALISCO UNIT PRICE",
            "TOTAL COST BASE BID",
            "ZAK DIRT UNIT PRICE",
            "TOTAL COST BASE BID",
            "FLATIRON UNIT PRICE",
            "TOTAL COST BASE BID",
            "ACC UNIT PRICE",
            "TOTAL COST BASE BID",
        ]
        for index, value in enumerate(headers, start=1):
            sheet.cell(1, index).value = value

        rows = [
            ["", "201-00000", "CLEARING AND GRUBBING", "LS", 1, 50, 50, 10, 10, 12, 12, 14, 14, 16, 16, 18, 18],
            ["", "202-00003", "REMOVAL OF STRUCTURE (SPECIAL)", "HR", 2, 20, 40, 20, 40, 24, 48, 28, 56, 32, 64, 36, 72],
            ["", "700", "F/A LANDSCAPE RESTORATION", "FA", 1, 10, 10, 50, 50, 60, 60, 70, 70, 80, 80, 90, 90],
        ]
        for row_index, row in enumerate(rows, start=2):
            for column_index, value in enumerate(row, start=1):
                sheet.cell(row_index, column_index).value = value

        summary = workbook.create_sheet("Sheet1")
        summary["A1"] = "IFB 21-09-02"
        summary["A2"] = "Arapahoe Rd Bridge at Big Dry Creek Project"
        summary["A3"] = "Bids"
        summary["A11"] = "Engr Est"
        summary["C11"] = 90

        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        tmp.close()
        workbook.save(tmp.name)
        return Path(tmp.name)


if __name__ == "__main__":
    unittest.main()
