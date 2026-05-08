from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path
from typing import Iterable

import openpyxl


SOURCE_PAGE_URL = "https://www.codot.gov/business/eema/itemcodebook"
SOURCE_ASSET_URL = "https://www.codot.gov/business/eema/assets/item-code-book-20251014.xlsx"
ITEM_CODE_PATTERN = re.compile(r"^\d{3}-\d{5}$")

KNOWN_SECTIONS = {
    "201": ("200", "Earthwork", "Clearing and Grubbing"),
    "202": ("200", "Earthwork", "Removal of Structures and Obstructions"),
    "203": ("200", "Earthwork", "Excavation and Embankment"),
    "206": ("200", "Earthwork", "Excavation and Backfill for Structures"),
    "207": ("200", "Earthwork", "Topsoil"),
    "208": ("200", "Earthwork", "Erosion Control"),
    "209": ("200", "Earthwork", "Watering and Dust Palliative"),
    "210": ("200", "Earthwork", "Reset Structures"),
    "212": ("200", "Earthwork", "Seeding, Fertilizing, Soil Conditioner and Sodding"),
    "213": ("200", "Earthwork", "Mulching"),
    "214": ("200", "Earthwork", "Planting"),
    "215": ("200", "Earthwork", "Transplanting"),
    "216": ("200", "Earthwork", "Soil Retention Covering"),
    "217": ("200", "Earthwork", "Herbicide Treatment"),
    "250": ("200", "Earthwork", "Environmental, Health and Safety Management"),
    "304": ("300", "Bases", "Aggregate Base Course"),
    "306": ("300", "Bases", "Reconditioning"),
    "307": ("300", "Bases", "Lime Treated Subgrade"),
    "401": ("400", "Pavements", "Plant Mix Pavements - General"),
    "403": ("400", "Pavements", "Hot Mix Asphalt"),
    "405": ("400", "Pavements", "Heating and Scarifying Treatment"),
    "406": ("400", "Pavements", "Cold Asphalt Pavement (Recycle)"),
    "407": ("400", "Pavements", "Prime Coat, Tack Coat and Rejuvenating Agent"),
    "408": ("400", "Pavements", "Joint and Crack Sealer"),
    "409": ("400", "Pavements", "Chip Seal"),
    "411": ("400", "Pavements", "Asphalt Materials"),
    "412": ("400", "Pavements", "Portland Cement Concrete Pavement"),
    "420": ("400", "Pavements", "Geosynthetics"),
    "501": ("500", "Structures", "Steel Sheet Piling"),
    "502": ("500", "Structures", "Piling"),
    "503": ("500", "Structures", "Drilled Shafts"),
    "504": ("500", "Structures", "Walls"),
    "506": ("500", "Structures", "Riprap"),
    "507": ("500", "Structures", "Slope and Ditch Paving"),
    "508": ("500", "Structures", "Timber Structures"),
    "509": ("500", "Structures", "Steel Structures"),
    "510": ("500", "Structures", "Structural Plate Structures"),
    "512": ("500", "Structures", "Bearing Device"),
    "514": ("500", "Structures", "Pedestrian and Bikeway Railing"),
    "515": ("500", "Structures", "Waterproofing Membrane"),
    "516": ("500", "Structures", "Dampproofing"),
    "517": ("500", "Structures", "Waterproofing"),
    "518": ("500", "Structures", "Waterstops and Expansion Joints"),
    "601": ("600", "Miscellaneous Construction", "Structural Concrete"),
    "602": ("600", "Miscellaneous Construction", "Reinforcing Steel"),
    "603": ("600", "Miscellaneous Construction", "Culverts and Sewers"),
    "604": ("600", "Miscellaneous Construction", "Manholes, Inlets and Meter Vaults"),
    "605": ("600", "Miscellaneous Construction", "Subsurface Drains"),
    "606": ("600", "Miscellaneous Construction", "Guardrail"),
    "607": ("600", "Miscellaneous Construction", "Fences"),
    "608": ("600", "Miscellaneous Construction", "Sidewalks and Bikeways"),
    "609": ("600", "Miscellaneous Construction", "Curb and Gutter"),
    "610": ("600", "Miscellaneous Construction", "Median Cover Material"),
    "611": ("600", "Miscellaneous Construction", "Cattle Guards"),
    "612": ("600", "Miscellaneous Construction", "Delineators and Reflectors"),
    "613": ("600", "Miscellaneous Construction", "Lighting"),
    "614": ("600", "Miscellaneous Construction", "Traffic Control Devices"),
    "615": ("600", "Miscellaneous Construction", "Water Control Devices"),
    "616": ("600", "Miscellaneous Construction", "Siphons"),
    "618": ("600", "Miscellaneous Construction", "Prestressed Concrete"),
    "619": ("600", "Miscellaneous Construction", "Water Lines"),
    "620": ("600", "Miscellaneous Construction", "Field Facilities"),
    "622": ("600", "Miscellaneous Construction", "Rest Areas and Buildings"),
    "623": ("600", "Miscellaneous Construction", "Irrigation System"),
    "624": ("600", "Miscellaneous Construction", "Drainage Pipe"),
    "625": ("600", "Miscellaneous Construction", "Construction Surveying"),
    "626": ("600", "Miscellaneous Construction", "Mobilization"),
    "627": ("600", "Miscellaneous Construction", "Pavement Marking"),
    "629": ("600", "Miscellaneous Construction", "Survey Monumentation"),
    "630": ("600", "Miscellaneous Construction", "Construction Zone Traffic Control"),
    "641": ("600", "Miscellaneous Construction", "Shotcrete"),
    "701": ("700", "Materials", "Hydraulic Cement"),
    "702": ("700", "Materials", "Bituminous Materials"),
    "703": ("700", "Materials", "Aggregates"),
    "704": ("700", "Materials", "Masonry Units"),
    "705": ("700", "Materials", "Joint, Waterproofing and Bearing Material"),
    "706": ("700", "Materials", "Concrete and Clay Pipe"),
    "707": ("700", "Materials", "Metal Pipe"),
    "708": ("700", "Materials", "Paints"),
    "709": ("700", "Materials", "Reinforcing Steel and Wire Rope"),
    "710": ("700", "Materials", "Fence and Guardrail"),
    "711": ("700", "Materials", "Concrete Curing Materials and Admixtures"),
    "712": ("700", "Materials", "Miscellaneous"),
    "713": ("700", "Materials", "Traffic Control Materials"),
    "714": ("700", "Materials", "Prestressed Unit Materials"),
    "715": ("700", "Materials", "Lighting and Electrical Materials"),
    "716": ("700", "Materials", "Water Line Materials"),
    "717": ("700", "Materials", "Rest Area and Building Materials"),
}


def clean_cell(value: object) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_header(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", "", clean_cell(value).lower())


def build_merged_value_lookup(worksheet) -> dict[tuple[int, int], object]:
    merged_values: dict[tuple[int, int], object] = {}
    for merged_range in worksheet.merged_cells.ranges:
        top_left = worksheet.cell(merged_range.min_row, merged_range.min_col).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for column in range(merged_range.min_col, merged_range.max_col + 1):
                merged_values[(row, column)] = top_left
    return merged_values


def row_values(worksheet, row_number: int, merged_values: dict[tuple[int, int], object]) -> list[str]:
    return [
        clean_cell(merged_values.get((row_number, column), worksheet.cell(row_number, column).value))
        for column in range(1, worksheet.max_column + 1)
    ]


def find_header_row(rows: list[list[str]]) -> tuple[int | None, dict[str, int]]:
    for row_index, row in enumerate(rows[:25]):
        normalized = [normalize_header(value) for value in row]
        code_index = find_any(normalized, {"itemcode", "itemnumber", "itemno", "item"})
        name_index = find_any(normalized, {"name", "longdescription", "description"})
        abbreviated_index = find_any(
            normalized,
            {"abbreviatedname", "abbreviateddescription", "shortdescription", "shortname", "abbrname"},
        )
        unit_index = find_any(normalized, {"unit", "units"})
        if None not in {code_index, name_index, abbreviated_index, unit_index}:
            return row_index, {
                "item_code": code_index,
                "official_description": name_index,
                "official_abbreviated_description": abbreviated_index,
                "official_unit": unit_index,
            }

    return None, {
        "item_code": 0,
        "official_description": 1,
        "official_abbreviated_description": 2,
        "official_unit": 3,
    }


def find_any(values: list[str], candidates: set[str]) -> int | None:
    for index, value in enumerate(values):
        if value in candidates:
            return index
    return None


def read_existing_canonical_mappings(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}

    with path.open(newline="", encoding="utf-8") as source:
        return {
            row["item_code"].strip().upper(): row.get("canonical_item_id", "").strip()
            for row in csv.DictReader(source)
            if row.get("item_code") and row.get("canonical_item_id", "").strip()
        }


def read_workbook_items(path: Path) -> tuple[list[dict[str, str]], list[str]]:
    workbook = openpyxl.load_workbook(path, read_only=False, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    merged_values = build_merged_value_lookup(worksheet)
    rows = [row_values(worksheet, row_number, merged_values) for row_number in range(1, worksheet.max_row + 1)]
    header_row_index, columns = find_header_row(rows)
    start_index = header_row_index + 1 if header_row_index is not None else 0
    items: list[dict[str, str]] = []
    rejected: list[str] = []
    seen_codes: set[str] = set()

    for row_number, row in enumerate(rows[start_index:], start=start_index + 1):
        if not row or all(not value for value in row):
            continue

        item_code = get_column(row, columns["item_code"]).upper()
        description = get_column(row, columns["official_description"])
        abbreviated = get_column(row, columns["official_abbreviated_description"])
        unit = get_column(row, columns["official_unit"]).upper()

        if not item_code and not description and not unit:
            continue

        if not item_code and description and not unit and items:
            items[-1]["official_description"] = f"{items[-1]['official_description']} {description}".strip()
            if abbreviated:
                items[-1]["official_abbreviated_description"] = (
                    f"{items[-1]['official_abbreviated_description']} {abbreviated}"
                ).strip()
            continue

        if not ITEM_CODE_PATTERN.fullmatch(item_code):
            rejected.append(f"Row {row_number}: invalid item code {item_code!r}")
            continue

        if not description or not unit:
            rejected.append(f"Row {row_number}: missing description or unit for {item_code}")
            continue

        if item_code in seen_codes:
            rejected.append(f"Row {row_number}: duplicate item code {item_code}")
            continue

        seen_codes.add(item_code)
        items.append(
            {
                "agency_item_id": f"co_cdot_{item_code}",
                "state": "CO",
                "agency": "CDOT",
                "item_code": item_code,
                "official_description": description,
                "official_abbreviated_description": abbreviated,
                "official_unit": unit,
                "canonical_item_id": "",
            }
        )

    return sorted(items, key=lambda item: item["item_code"]), rejected


def get_column(row: list[str], index: int) -> str:
    return row[index] if index < len(row) else ""


def apply_existing_mappings(items: list[dict[str, str]], mappings: dict[str, str]) -> None:
    for item in items:
        item["canonical_item_id"] = mappings.get(item["item_code"], "")


def build_spec_sections(items: Iterable[dict[str, str]], source_year: int) -> list[dict[str, str]]:
    prefixes = sorted({item["item_code"][:3] for item in items})
    rows: list[dict[str, str]] = []

    for prefix in prefixes:
        if prefix in KNOWN_SECTIONS:
            division_prefix, division_title, section_title = KNOWN_SECTIONS[prefix]
        else:
            division_prefix = f"{prefix[0]}00"
            division_title = "Other CDOT item prefixes"
            section_title = f"Prefix {prefix}"

        rows.append(
            {
                "section_prefix": prefix,
                "division_prefix": division_prefix,
                "division_title": division_title,
                "section_title": section_title,
                "source_year": str(source_year),
                "source_url": SOURCE_PAGE_URL,
            }
        )

    return rows


def validate_outputs(items: list[dict[str, str]], sections: list[dict[str, str]]) -> list[str]:
    issues: list[str] = []
    codes = [item["item_code"] for item in items]
    duplicate_codes = sorted({code for code in codes if codes.count(code) > 1})
    if duplicate_codes:
        issues.append(f"Duplicate item codes: {', '.join(duplicate_codes[:10])}")

    required_item_fields = [
        "agency_item_id",
        "state",
        "agency",
        "item_code",
        "official_description",
        "official_unit",
    ]
    for item in items:
        missing = [field for field in required_item_fields if not item.get(field)]
        if missing:
            issues.append(f"{item.get('item_code', '<missing code>')}: missing {', '.join(missing)}")

    section_prefixes = {section["section_prefix"] for section in sections}
    missing_sections = sorted({item["item_code"][:3] for item in items} - section_prefixes)
    if missing_sections:
        issues.append(f"Missing section rows for prefixes: {', '.join(missing_sections)}")

    return issues


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Import the CDOT Item Code Book workbook into static app CSV files."
    )
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--agency-items", default=Path("public/data/agency_items.csv"), type=Path)
    parser.add_argument("--spec-sections", default=Path("public/data/spec_sections.csv"), type=Path)
    parser.add_argument("--source-year", default=2026, type=int)
    parser.add_argument("--min-row-count", default=4500, type=int)
    args = parser.parse_args()

    mappings = read_existing_canonical_mappings(args.agency_items)
    items, rejected = read_workbook_items(args.source)
    apply_existing_mappings(items, mappings)
    sections = build_spec_sections(items, args.source_year)
    issues = validate_outputs(items, sections)

    if len(items) < args.min_row_count:
        issues.append(f"Expected at least {args.min_row_count} item rows, found {len(items)}")

    if issues:
        for issue in issues:
            print(f"ERROR: {issue}")
        raise SystemExit(1)

    write_csv(
        args.agency_items,
        items,
        [
            "agency_item_id",
            "state",
            "agency",
            "item_code",
            "official_description",
            "official_abbreviated_description",
            "official_unit",
            "canonical_item_id",
        ],
    )
    write_csv(
        args.spec_sections,
        sections,
        [
            "section_prefix",
            "division_prefix",
            "division_title",
            "section_title",
            "source_year",
            "source_url",
        ],
    )

    mapped_count = sum(1 for item in items if item["canonical_item_id"])
    print(f"Wrote {len(items)} agency item rows.")
    print(f"Wrote {len(sections)} spec section rows.")
    print(f"Preserved {mapped_count} canonical item mappings.")
    if rejected:
        print(f"Rejected {len(rejected)} workbook row(s).")
        for message in rejected[:20]:
            print(f"WARNING: {message}")


if __name__ == "__main__":
    main()
