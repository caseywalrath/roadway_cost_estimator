from __future__ import annotations

import argparse
import csv
import re
from collections import defaultdict
from pathlib import Path

import openpyxl


SOURCE_URL = (
    "https://www.codot.gov/business/designsupport/cdot-construction-specifications/"
    "2025-construction-specifications/specs-book"
)

SECTIONS = {
    "201": ("200", "Earthwork", "Clearing and Grubbing"),
    "202": ("200", "Earthwork", "Removal of Structures and Obstructions"),
    "203": ("200", "Earthwork", "Excavation and Embankment"),
    "206": ("200", "Earthwork", "Excavation and Backfill for Structures"),
    "207": ("200", "Earthwork", "Topsoil"),
    "208": ("200", "Earthwork", "Erosion Control"),
    "209": ("200", "Earthwork", "Watering and Dust Palliative"),
    "210": ("200", "Earthwork", "Reset Structures"),
    "212": ("200", "Earthwork", "Seeding, Fertilizing, Soil Conditioner and Sodding"),
    "213": ("200", "Earthwork", "Mulching"),
    "214": ("200", "Earthwork", "Planting"),
    "215": ("200", "Earthwork", "Transplanting"),
    "216": ("200", "Earthwork", "Soil Retention Covering"),
    "217": ("200", "Earthwork", "Herbicide Treatment"),
    "250": ("200", "Earthwork", "Environmental, Health and Safety Management"),
    "304": ("300", "Bases", "Aggregate Base Course"),
    "306": ("300", "Bases", "Reconditioning"),
    "307": ("300", "Bases", "Lime Treated Subgrade"),
    "401": ("400", "Pavements", "Plant Mix Pavements - General"),
    "403": ("400", "Pavements", "Hot Mix Asphalt"),
    "405": ("400", "Pavements", "Heating and Scarifying Treatment"),
    "406": ("400", "Pavements", "Cold Asphalt Pavement (Recycle)"),
    "407": ("400", "Pavements", "Prime Coat, Tack Coat and Rejuvenating Agent"),
    "408": ("400", "Pavements", "Joint and Crack Sealer"),
    "409": ("400", "Pavements", "Chip Seal"),
    "411": ("400", "Pavements", "Asphalt Materials"),
    "412": ("400", "Pavements", "Portland Cement Concrete Pavement"),
    "420": ("400", "Pavements", "Geosynthetics"),
    "501": ("500", "Structures", "Steel Sheet Piling"),
    "502": ("500", "Structures", "Piling"),
    "503": ("500", "Structures", "Drilled Shafts"),
    "504": ("500", "Structures", "Walls"),
    "506": ("500", "Structures", "Riprap"),
    "507": ("500", "Structures", "Slope and Ditch Paving"),
    "508": ("500", "Structures", "Timber Structures"),
    "509": ("500", "Structures", "Steel Structures"),
    "510": ("500", "Structures", "Structural Plate Structures"),
    "512": ("500", "Structures", "Bearing Device"),
    "514": ("500", "Structures", "Pedestrian and Bikeway Railing"),
    "515": ("500", "Structures", "Waterproofing Membrane"),
    "516": ("500", "Structures", "Dampproofing"),
    "517": ("500", "Structures", "Waterproofing"),
    "518": ("500", "Structures", "Waterstops and Expansion Joints"),
    "601": ("600", "Miscellaneous Construction", "Structural Concrete"),
    "602": ("600", "Miscellaneous Construction", "Reinforcing Steel"),
    "603": ("600", "Miscellaneous Construction", "Culverts and Sewers"),
    "604": ("600", "Miscellaneous Construction", "Manholes, Inlets and Meter Vaults"),
    "605": ("600", "Miscellaneous Construction", "Subsurface Drains"),
    "606": ("600", "Miscellaneous Construction", "Guardrail"),
    "607": ("600", "Miscellaneous Construction", "Fences"),
    "608": ("600", "Miscellaneous Construction", "Sidewalks and Bikeways"),
    "609": ("600", "Miscellaneous Construction", "Curb and Gutter"),
    "610": ("600", "Miscellaneous Construction", "Median Cover Material"),
    "611": ("600", "Miscellaneous Construction", "Cattle Guards"),
    "612": ("600", "Miscellaneous Construction", "Delineators and Reflectors"),
    "613": ("600", "Miscellaneous Construction", "Lighting"),
    "614": ("600", "Miscellaneous Construction", "Traffic Control Devices"),
    "615": ("600", "Miscellaneous Construction", "Water Control Devices"),
    "616": ("600", "Miscellaneous Construction", "Siphons"),
    "618": ("600", "Miscellaneous Construction", "Prestressed Concrete"),
    "619": ("600", "Miscellaneous Construction", "Water Lines"),
    "620": ("600", "Miscellaneous Construction", "Field Facilities"),
    "622": ("600", "Miscellaneous Construction", "Rest Areas and Buildings"),
    "623": ("600", "Miscellaneous Construction", "Irrigation System"),
    "624": ("600", "Miscellaneous Construction", "Drainage Pipe"),
    "625": ("600", "Miscellaneous Construction", "Construction Surveying"),
    "626": ("600", "Miscellaneous Construction", "Mobilization"),
    "627": ("600", "Miscellaneous Construction", "Pavement Marking"),
    "629": ("600", "Miscellaneous Construction", "Survey Monumentation"),
    "630": ("600", "Miscellaneous Construction", "Construction Zone Traffic Control"),
    "641": ("600", "Miscellaneous Construction", "Shotcrete"),
    "701": ("700", "Materials", "Hydraulic Cement"),
    "702": ("700", "Materials", "Bituminous Materials"),
    "703": ("700", "Materials", "Aggregates"),
    "704": ("700", "Materials", "Masonry Units"),
    "705": ("700", "Materials", "Joint, Waterproofing and Bearing Material"),
    "706": ("700", "Materials", "Concrete and Clay Pipe"),
    "707": ("700", "Materials", "Metal Pipe"),
    "708": ("700", "Materials", "Paints"),
    "709": ("700", "Materials", "Reinforcing Steel and Wire Rope"),
    "710": ("700", "Materials", "Fence and Guardrail"),
    "711": ("700", "Materials", "Concrete Curing Materials and Admixtures"),
    "712": ("700", "Materials", "Miscellaneous"),
    "713": ("700", "Materials", "Traffic Control Materials"),
    "714": ("700", "Materials", "Prestressed Unit Materials"),
    "715": ("700", "Materials", "Lighting and Electrical Materials"),
    "716": ("700", "Materials", "Water Line Materials"),
    "717": ("700", "Materials", "Rest Area and Building Materials"),
}


def read_existing_rows(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}

    with path.open(newline="", encoding="utf-8") as source:
        return {
            row["item_code"].strip().upper(): {
                "agency_item_id": row["agency_item_id"].strip(),
                "state": row["state"].strip(),
                "agency": row["agency"].strip(),
                "item_code": row["item_code"].strip().upper(),
                "official_description": row["official_description"].strip(),
                "official_unit": row["official_unit"].strip().upper(),
                "canonical_item_id": row["canonical_item_id"].strip(),
            }
            for row in csv.DictReader(source)
            if row.get("item_code") and row.get("canonical_item_id", "").strip()
        }


def read_workbook_items(path: Path) -> list[dict[str, str]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    items: list[dict[str, str]] = []

    for row in worksheet.iter_rows(values_only=True):
        code = str(row[0] or "").strip().upper()
        if not re.fullmatch(r"\d{3}-\d{5}", code):
            continue

        prefix = code[:3]
        if prefix not in SECTIONS:
            continue

        long_description = str(row[1] or "").strip()
        unit = str(row[3] or "").strip().upper()
        if not long_description or not unit:
            continue

        items.append(
            {
                "agency_item_id": f"co_cdot_{code}",
                "state": "CO",
                "agency": "CDOT",
                "item_code": code,
                "official_description": long_description,
                "official_unit": unit,
                "canonical_item_id": "",
            }
        )

    return items


def sample_items(
    items: list[dict[str, str]],
    target_count: int,
    existing_by_code: dict[str, dict[str, str]],
) -> list[dict[str, str]]:
    by_prefix: dict[str, list[dict[str, str]]] = defaultdict(list)
    by_code = {item["item_code"]: item for item in items}

    for item in items:
        by_prefix[item["item_code"][:3]].append(item)

    selected_codes: set[str] = set()
    selected: list[dict[str, str]] = []

    def add_item(item: dict[str, str]) -> None:
        code = item["item_code"]
        if code in selected_codes:
            return
        selected_codes.add(code)
        selected.append(item.copy())

    for code in sorted(existing_by_code):
        item = by_code.get(code)
        if item:
            add_item(item)

    for prefix in sorted(SECTIONS):
        for item in by_prefix.get(prefix, [])[:2]:
            add_item(item)

    prefix_order = sorted(
        (prefix for prefix in SECTIONS if by_prefix.get(prefix)),
        key=lambda prefix: (-len(by_prefix[prefix]), prefix),
    )
    index_by_prefix = {prefix: 2 for prefix in prefix_order}

    while len(selected) < target_count:
        made_progress = False
        for prefix in prefix_order:
            index = index_by_prefix[prefix]
            candidates = by_prefix[prefix]
            if index >= len(candidates):
                continue
            add_item(candidates[index])
            index_by_prefix[prefix] = index + 1
            made_progress = True
            if len(selected) >= target_count:
                break
        if not made_progress:
            break

    selected = sorted(selected[:target_count], key=lambda item: item["item_code"])

    for item in selected:
        existing = existing_by_code.get(item["item_code"])
        if existing:
            item.update(existing)

    return selected


def write_agency_items(path: Path, rows: list[dict[str, str]]) -> None:
    fieldnames = [
        "agency_item_id",
        "state",
        "agency",
        "item_code",
        "official_description",
        "official_unit",
        "canonical_item_id",
    ]
    with path.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_spec_sections(path: Path, selected_items: list[dict[str, str]]) -> None:
    selected_prefixes = sorted({item["item_code"][:3] for item in selected_items})
    fieldnames = [
        "section_prefix",
        "division_prefix",
        "division_title",
        "section_title",
        "source_year",
        "source_url",
    ]
    with path.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=fieldnames)
        writer.writeheader()
        for prefix in selected_prefixes:
            division_prefix, division_title, section_title = SECTIONS[prefix]
            writer.writerow(
                {
                    "section_prefix": prefix,
                    "division_prefix": division_prefix,
                    "division_title": division_title,
                    "section_title": section_title,
                    "source_year": "2025",
                    "source_url": SOURCE_URL,
                }
            )


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", required=True, type=Path)
    parser.add_argument("--agency-items", default=Path("public/data/agency_items.csv"), type=Path)
    parser.add_argument("--spec-sections", default=Path("public/data/spec_sections.csv"), type=Path)
    parser.add_argument("--target-count", default=200, type=int)
    args = parser.parse_args()

    existing_by_code = read_existing_rows(args.agency_items)
    workbook_items = read_workbook_items(args.source)
    sampled_items = sample_items(workbook_items, args.target_count, existing_by_code)

    write_agency_items(args.agency_items, sampled_items)
    write_spec_sections(args.spec_sections, sampled_items)

    print(f"Wrote {len(sampled_items)} agency item rows.")
    print(f"Wrote {len({item['item_code'][:3] for item in sampled_items})} spec section rows.")


if __name__ == "__main__":
    main()
