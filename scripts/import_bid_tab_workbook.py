from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable


DEFAULT_WORKBOOK = Path(r"C:\Users\Casey.Walrath\Downloads\2021 12 16 Bid Tabs Watson Ave Roundabout.xlsx")
DEFAULT_SOURCE_ID = "fhu_bid_tab_watson_ave_roundabout_2021_12_16"
DEFAULT_SOURCE_LABEL = "FHU Civil Group Bid Tabs - Watson Avenue Roundabout"
DEFAULT_SOURCE_TYPE = "public_bid_tab"
DEFAULT_SOURCE_YEAR = "2021"
DEFAULT_DATE_BASIS = "2021-12-16"
DEFAULT_PROJECT_ID = "fhu_bid_tab_watson_ave_roundabout_2021_12_16_320150"
DEFAULT_ROW_PREFIX = "fhu_watson_20211216"

DEFAULT_SOURCES = Path("public/data/sources.csv")
DEFAULT_PROJECTS = Path("public/data/projects.csv")
DEFAULT_OBSERVATIONS = Path("public/data/item_observations.csv")
DEFAULT_BIDDER_BIDS = Path("public/data/bidder_bids.csv")
DEFAULT_BIDDER_ITEMS = Path("public/data/bidder_item_observations.csv")
DEFAULT_BID_TAB_ITEMS = Path("public/data/bid_tab_items.csv")
DEFAULT_AGENCY_ITEMS = Path("public/data/agency_items.csv")
DEFAULT_STAGING_ITEMS = Path("public/data/imports/fhu_bid_tab_watson_ave_roundabout_2021_12_16_item_unit_costs.csv")
DEFAULT_STAGING_BIDDER_BIDS = Path("public/data/imports/fhu_bid_tab_watson_ave_roundabout_2021_12_16_bidder_bids.csv")
DEFAULT_STAGING_BIDDER_ITEMS = Path("public/data/imports/fhu_bid_tab_watson_ave_roundabout_2021_12_16_bidder_item_observations.csv")
DEFAULT_STAGING_MATCH_CANDIDATES = Path("public/data/imports/fhu_bid_tab_watson_ave_roundabout_2021_12_16_match_candidates.csv")
DEFAULT_MASTER_CONFIG = Path("data/staging/co/cost_estimate_master_sources.json")

ITEM_CODE_PATTERN = re.compile(r"^\d{3}-\d{5}$")
BID_FORM_ITEM_CODE_PATTERN = re.compile(r"^(?:\d{3}-\d{5}|\d{3})$")
PROJECT_NAME_PATTERN = re.compile(r"PROJECT NAME:\s*(?P<value>.+)", re.IGNORECASE)
PROJECT_NUMBER_PATTERN = re.compile(r"PROJECT NO\.:\s*(?P<value>.+)", re.IGNORECASE)

SOURCE_FIELDS = [
    "source_id", "source_type", "agency", "state", "source_label", "source_date", "data_year",
    "source_url", "source_file_name", "sha256", "parser_name", "parser_version", "notes",
]
PROJECT_FIELDS = [
    "project_id",
    "project_name",
    "agency_owner",
    "state",
    "county_region",
    "work_type",
    "estimate_let_date",
    "source_id",
    "project_number",
    "project_location_raw",
    "contractor",
    "district",
    "terrain",
    "bid_count",
    "awarded_bid_total",
    "award_index",
]
OBSERVATION_FIELDS = [
    "observation_id",
    "project_id",
    "source_id",
    "agency_item_code",
    "description_raw",
    "description_normalized",
    "unit_raw",
    "unit_normalized",
    "quantity",
    "unit_price",
    "extended_price",
    "discipline",
    "price_type",
    "date_basis",
]
BIDDER_BID_FIELDS = [
    "bid_id",
    "project_id",
    "source_id",
    "bidder_name",
    "bid_total",
    "bid_rank",
    "apparent_low",
    "is_awarded",
]
BIDDER_ITEM_FIELDS = [
    "bidder_item_observation_id",
    "bid_tab_item_id",
    "bid_id",
    "project_id",
    "source_id",
    "agency_item_code",
    "description_raw",
    "unit_raw",
    "unit_normalized",
    "quantity",
    "unit_price",
    "extended_price",
]
BID_TAB_ITEM_FIELDS = [
    "bid_tab_item_id",
    "project_id",
    "source_id",
    "source_file",
    "sheet_name",
    "workbook_row",
    "project_number",
    "source_item_number",
    "source_item_code",
    "source_item_code_system",
    "source_spec_raw",
    "source_item_description",
    "item_code",
    "item_description",
    "unit_raw",
    "unit_normalized",
    "quantity",
    "engineer_estimate_unit_price",
    "average_bid_unit_price",
    "matched_agency_item_code",
    "match_status",
    "date_basis",
]
STAGING_ITEM_FIELDS = BID_TAB_ITEM_FIELDS + [
    "engineer_estimate_quantity",
    "audit_estimate_unit_price",
    "audit_estimate_extended_price",
]
MATCH_CANDIDATE_FIELDS = [
    "bid_tab_item_id",
    "source_item_code",
    "source_item_description",
    "unit_normalized",
    "candidate_agency_item_code",
    "candidate_description",
    "candidate_unit",
    "candidate_count",
    "suggestion_status",
]


@dataclass(frozen=True)
class PriceColumnGroup:
    name: str
    unit_col: int
    total_col: int
    role: str


@dataclass(frozen=True)
class ParsedBidTab:
    project_name: str
    project_number: str
    item_rows: list[dict[str, str]]
    bidder_bids: list[dict[str, str]]
    bidder_items: list[dict[str, str]]
    observations: list[dict[str, str]]
    warnings: list[str]


def normalize_description(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", value.lower().replace("&", " and "))).strip()


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "_", value.lower()).strip("_")


def decimal_value(value: object, default: str | None = None) -> Decimal:
    if value is None and default is not None:
        value = default
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value)).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    if isinstance(value, str):
        cleaned = value.replace("$", "").replace(",", "").strip()
        if not cleaned and default is not None:
            cleaned = default
        try:
            return Decimal(cleaned).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        except InvalidOperation as error:
            raise ValueError(f"Expected numeric value, received {value!r}") from error
    raise ValueError(f"Expected numeric value, received {value!r}")


def decimal_text(value: Decimal) -> str:
    return str(value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))


def quantity_text(value: Decimal) -> str:
    return format(value.normalize(), "f")


def normalize_unit(value: object) -> str:
    raw = str(value or "").strip().upper()
    compact = re.sub(r"[^A-Z0-9]+", " ", raw).strip()
    mapping = {
        "EA": "EACH",
        "EACH": "EACH",
        "LS": "L S",
        "L S": "L S",
        "LUMP SUM": "L S",
        "LB": "LB",
        "POUND": "LB",
        "POUNDS": "LB",
        "SF": "SF",
        "SQ FT": "SF",
        "SQUARE FOOT": "SF",
        "SY": "SY",
        "SQ YD": "SY",
        "SQUARE YARD": "SY",
        "CY": "CY",
        "CU YD": "CY",
        "CUBIC YARD": "CY",
        "LF": "LF",
        "FOOT": "LF",
        "FEET": "LF",
        "AC": "ACRE",
        "ACRE": "ACRE",
        "HR": "HOUR",
        "HOUR": "HOUR",
        "HOURS": "HOUR",
        "DY": "DAY",
        "DAY": "DAY",
        "DAYS": "DAY",
        "FA": "F A",
        "F A": "F A",
        "F/A": "F A",
        "F A": "F A",
    }
    return mapping.get(compact, raw)


def read_csv(path: Path) -> list[dict[str, str]]:
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8") as source:
        return [dict(row) for row in csv.DictReader(source)]


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as output:
        writer = csv.DictWriter(output, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def normalize_existing_rows(rows: Iterable[dict[str, str]], fieldnames: list[str]) -> list[dict[str, str]]:
    return [{field: row.get(field, "") for field in fieldnames} for row in rows]


def agency_item_codes(path: Path) -> set[str]:
    return {row["item_code"].strip().upper() for row in read_csv(path) if row.get("item_code")}


def agency_item_rows(path: Path) -> list[dict[str, str]]:
    return read_csv(path)


def bid_tab_item_row(
    bid_tab_item_id: str,
    project_id: str,
    source_id: str,
    workbook_path: Path,
    sheet_name: str,
    workbook_row: int,
    project_number: str,
    source_item_number: str,
    source_item_code: str,
    source_item_code_system: str,
    source_spec_raw: str,
    source_item_description: str,
    item_code: str,
    unit_raw: str,
    unit_normalized: str,
    quantity: Decimal,
    engineer_unit_price: Decimal,
    average_unit_price: Decimal,
    date_basis: str,
    matched_agency_item_code: str | None = None,
    match_status: str | None = None,
) -> dict[str, str]:
    normalized_item_code = item_code.strip().upper()
    matched_code = matched_agency_item_code
    if matched_code is None:
        matched_code = normalized_item_code if ITEM_CODE_PATTERN.match(normalized_item_code) else ""
    status = match_status
    if status is None:
        status = "matched" if matched_code else "source_cdot_prefix_only" if source_item_code_system == "CDOT" else "unmatched"

    return {
        "bid_tab_item_id": bid_tab_item_id,
        "project_id": project_id,
        "source_id": source_id,
        "source_file": workbook_path.name,
        "sheet_name": sheet_name,
        "workbook_row": str(workbook_row),
        "project_number": project_number,
        "source_item_number": source_item_number,
        "source_item_code": source_item_code,
        "source_item_code_system": source_item_code_system,
        "source_spec_raw": source_spec_raw,
        "source_item_description": source_item_description,
        "item_code": normalized_item_code,
        "item_description": source_item_description,
        "unit_raw": unit_raw,
        "unit_normalized": unit_normalized,
        "quantity": decimal_text(quantity),
        "engineer_estimate_unit_price": decimal_text(engineer_unit_price),
        "average_bid_unit_price": decimal_text(average_unit_price),
        "matched_agency_item_code": matched_code,
        "match_status": status,
        "date_basis": date_basis,
    }


def source_code_system(value: str) -> str:
    upper_value = value.strip().upper()
    if upper_value.startswith("COA"):
        return "COA"
    if upper_value.startswith("CDOT") or ITEM_CODE_PATTERN.match(upper_value) or re.match(r"^\d{3}$", upper_value):
        return "CDOT"
    if not upper_value:
        return "SOURCE_ONLY"
    return "SOURCE"


def parse_workbook(
    workbook_path: Path,
    source_id: str,
    project_id: str,
    row_prefix: str,
    date_basis: str,
) -> ParsedBidTab:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise SystemExit("Missing dependency: openpyxl. Run with the bundled Codex Python runtime.") from error

    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    sheet = workbook[workbook.sheetnames[0]]
    formula_workbook = load_workbook(workbook_path, data_only=False, read_only=False)
    formula_sheet = formula_workbook[sheet.title]

    results_sheet = next((candidate for candidate in workbook.worksheets if find_results_header_row(candidate)), None)
    if results_sheet:
        return parse_results_workbook(
            workbook_path,
            results_sheet,
            formula_workbook[results_sheet.title],
            source_id,
            project_id,
            row_prefix,
            date_basis,
        )

    split_header = find_split_bid_tab_header_rows(sheet)
    if split_header:
        return parse_split_header_bid_tab_workbook(
            workbook_path,
            sheet,
            formula_sheet,
            source_id,
            project_id,
            row_prefix,
            date_basis,
            split_header[0],
            split_header[1],
        )

    bid_form_header = find_bid_form_header_row(sheet)
    if bid_form_header:
        return parse_bid_form_workbook(
            workbook_path,
            workbook,
            formula_workbook,
            source_id,
            project_id,
            row_prefix,
            date_basis,
            bid_form_header,
        )

    return parse_saq_workbook(
        workbook_path,
        sheet,
        formula_sheet,
        source_id,
        project_id,
        row_prefix,
        date_basis,
    )


def parse_results_workbook(
    workbook_path: Path,
    sheet,
    formula_sheet,
    source_id: str,
    project_id: str,
    row_prefix: str,
    date_basis: str,
) -> ParsedBidTab:
    header_row = find_results_header_row(sheet)
    if not header_row:
        raise ValueError("Results workbook must include No., Spec*, Item, Unit, and Quantity headers.")

    column_by_header = results_column_map(sheet, header_row)
    groups = detect_results_price_groups(sheet, header_row - 1, header_row)
    engineer_group = next((group for group in groups if group.role == "engineer"), None)
    bidder_groups = [group for group in groups if group.role == "bidder"]

    if not engineer_group or not bidder_groups:
        raise ValueError("Results workbook must include engineer estimate and bidder Unit Cost / Extended Cost pairs.")

    item_rows: list[dict[str, str]] = []
    observations: list[dict[str, str]] = []
    bidder_items: list[dict[str, str]] = []
    warnings: list[str] = []
    bid_totals: dict[str, Decimal] = {group.name: Decimal("0.00") for group in bidder_groups}
    engineer_total = Decimal("0.00")

    project_name = "Ralston Road - Yukon Street to Garrison Street"
    project_number = "18-ST-40"
    total_row = find_results_total_row(sheet, column_by_header["ITEM"], header_row + 1)

    for row_index in range(header_row + 1, total_row):
        source_item_value = sheet.cell(row_index, column_by_header["NO."]).value
        if source_item_value is None or str(source_item_value).strip() == "":
            continue

        source_item_number = normalize_source_item_number(source_item_value)
        source_spec_raw = str(sheet.cell(row_index, column_by_header["SPEC*"]).value or "").strip()
        description = str(sheet.cell(row_index, column_by_header["ITEM"]).value or "").strip()
        unit_raw = str(sheet.cell(row_index, column_by_header["UNIT"]).value or "").strip()
        unit_normalized = normalize_unit(unit_raw)
        quantity = decimal_value(sheet.cell(row_index, column_by_header["QUANTITY"]).value)
        engineer_unit_price = decimal_value(sheet.cell(row_index, engineer_group.unit_col).value)
        engineer_extended_price = (quantity * engineer_unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        engineer_total += engineer_extended_price

        bidder_unit_prices: list[Decimal] = []
        row_id = f"{row_prefix}_row_{row_index:04d}"
        source_item_code = source_spec_raw or f"No. {source_item_number}"
        code_system = source_code_system(source_item_code)
        reconciliation = results_reconciliation(sheet, column_by_header, row_index)
        has_reconciliation = "CDOT ITEM CODE" in column_by_header
        matched_item_code = reconciliation["item_code"]
        matched_description = reconciliation["description"] or description
        matched_unit_raw = reconciliation["unit_raw"] or unit_raw
        matched_unit_normalized = normalize_unit(matched_unit_raw)
        match_status = (
            "matched"
            if matched_item_code
            else "unmatched"
            if has_reconciliation
            else "source_cdot_prefix_only"
            if code_system == "CDOT"
            else "unmatched"
        )

        item_rows.append(
            bid_tab_item_row(
                row_id,
                project_id,
                source_id,
                workbook_path,
                sheet.title,
                row_index,
                project_number,
                source_item_number,
                source_item_code,
                code_system,
                source_spec_raw,
                description,
                matched_item_code,
                unit_raw,
                unit_normalized,
                quantity,
                engineer_unit_price,
                Decimal("0.00"),
                date_basis,
                matched_item_code,
                match_status,
            )
        )

        for bidder_group in bidder_groups:
            unit_price = decimal_value(sheet.cell(row_index, bidder_group.unit_col).value)
            extended_price = (quantity * unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            bid_id = f"{row_prefix}_{slugify(bidder_group.name)}"
            bidder_unit_prices.append(unit_price)
            bid_totals[bidder_group.name] += extended_price
            bidder_items.append(
                {
                    "bidder_item_observation_id": f"{row_id}_{slugify(bidder_group.name)}",
                    "bid_tab_item_id": row_id,
                    "bid_id": bid_id,
                    "project_id": project_id,
                    "source_id": source_id,
                    "agency_item_code": matched_item_code,
                    "description_raw": matched_description if matched_item_code else description,
                    "unit_raw": matched_unit_raw if matched_item_code else unit_raw,
                    "unit_normalized": matched_unit_normalized if matched_item_code else unit_normalized,
                    "quantity": decimal_text(quantity),
                    "unit_price": decimal_text(unit_price),
                    "extended_price": decimal_text(extended_price),
                }
            )
            validate_line_total(
                sheet.cell(row_index, bidder_group.total_col).value,
                formula_sheet.cell(row_index, bidder_group.total_col).value,
                extended_price,
                warnings,
                row_index,
                bidder_group.name,
            )

        validate_line_total(
            sheet.cell(row_index, engineer_group.total_col).value,
            formula_sheet.cell(row_index, engineer_group.total_col).value,
            engineer_extended_price,
            warnings,
            row_index,
            "Engineer estimate",
        )

        average_unit_price = (sum(bidder_unit_prices) / Decimal(len(bidder_unit_prices))).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        item_rows[-1]["average_bid_unit_price"] = decimal_text(average_unit_price)

        if matched_item_code:
            observations.extend(
                [
                    promoted_observation(
                        f"{row_id}_engineer_estimate",
                        project_id,
                        source_id,
                        matched_item_code,
                        matched_description,
                        matched_unit_raw,
                        matched_unit_normalized,
                        quantity,
                        engineer_unit_price,
                        "public_bid_tab_engineer_estimate",
                        date_basis,
                    ),
                    promoted_observation(
                        f"{row_id}_average",
                        project_id,
                        source_id,
                        matched_item_code,
                        matched_description,
                        matched_unit_raw,
                        matched_unit_normalized,
                        quantity,
                        average_unit_price,
                        "public_bid_tab_average",
                        date_basis,
                    ),
                ]
            )

    published_engineer_total = decimal_value(sheet.cell(total_row, engineer_group.total_col).value)
    if abs(published_engineer_total - engineer_total) > Decimal("0.01"):
        warnings.append(
            f"Results total engineer estimate {published_engineer_total} differs from itemized engineer estimate total {engineer_total}"
        )

    for bidder_group in bidder_groups:
        published_total = decimal_value(sheet.cell(total_row, bidder_group.total_col).value)
        calculated_total = bid_totals[bidder_group.name]
        if abs(published_total - calculated_total) > Decimal("0.01"):
            warnings.append(
                f"Results total {bidder_group.name} {published_total} differs from calculated total {calculated_total}"
            )

    ranked_bidders = sorted(bid_totals.items(), key=lambda item: item[1])
    bidder_bids = [
        {
            "bid_id": f"{row_prefix}_{slugify(name)}",
            "project_id": project_id,
            "source_id": source_id,
            "bidder_name": name,
            "bid_total": decimal_text(total),
            "bid_rank": str(index),
            "apparent_low": "true" if index == 1 else "false",
        }
        for index, (name, total) in enumerate(ranked_bidders, start=1)
    ]

    return ParsedBidTab(project_name, project_number, item_rows, bidder_bids, bidder_items, observations, warnings)


def parse_saq_workbook(
    workbook_path: Path,
    sheet,
    formula_sheet,
    source_id: str,
    project_id: str,
    row_prefix: str,
    date_basis: str,
) -> ParsedBidTab:
    project_name = ""
    project_number = ""
    for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20)):
        for cell in row:
            if not isinstance(cell.value, str):
                continue
            name_match = PROJECT_NAME_PATTERN.search(cell.value)
            number_match = PROJECT_NUMBER_PATTERN.search(cell.value)
            if name_match:
                project_name = name_match.group("value").strip().title()
            if number_match:
                project_number = number_match.group("value").strip()

    header_row = find_header_row(sheet)
    group_row = header_row - 1
    groups = detect_price_groups(sheet, group_row, header_row)
    engineer_group = next((group for group in groups if group.role == "engineer"), None)
    average_group = next((group for group in groups if group.role == "average"), None)
    bidder_groups = [group for group in groups if group.role == "bidder"]

    if not project_name or not project_number:
        raise ValueError("Workbook header must include PROJECT NAME and PROJECT NO.")
    if not engineer_group or not average_group or not bidder_groups:
        raise ValueError("Workbook must include engineer estimate, at least one bidder, and average bid columns.")

    item_rows: list[dict[str, str]] = []
    observations: list[dict[str, str]] = []
    bidder_items: list[dict[str, str]] = []
    bid_totals: dict[str, Decimal] = {group.name: Decimal("0.00") for group in bidder_groups}
    warnings: list[str] = []

    for row_index in range(header_row + 2, sheet.max_row + 1):
        item_code = sheet.cell(row_index, 3).value
        if not isinstance(item_code, str) or not ITEM_CODE_PATTERN.match(item_code.strip()):
            continue

        item_code = item_code.strip().upper()
        description = str(sheet.cell(row_index, 4).value or "").strip()
        quantity = decimal_value(sheet.cell(row_index, 5).value)
        unit_raw = str(sheet.cell(row_index, 6).value or "").strip()
        unit_normalized = normalize_unit(unit_raw)
        engineer_unit_price = decimal_value(sheet.cell(row_index, engineer_group.unit_col).value)
        average_unit_price = decimal_value(sheet.cell(row_index, average_group.unit_col).value)

        row_id = f"{row_prefix}_row_{row_index:04d}"
        item_rows.append(
            bid_tab_item_row(
                row_id,
                project_id,
                source_id,
                workbook_path,
                sheet.title,
                row_index,
                project_number,
                "",
                item_code,
                "CDOT",
                item_code,
                description,
                item_code,
                unit_raw,
                unit_normalized,
                quantity,
                engineer_unit_price,
                average_unit_price,
                date_basis,
            )
        )

        observations.extend(
            [
                promoted_observation(
                    f"{row_id}_engineer_estimate",
                    project_id,
                    source_id,
                    item_code,
                    description,
                    unit_raw,
                    unit_normalized,
                    quantity,
                    engineer_unit_price,
                    "public_bid_tab_engineer_estimate",
                    date_basis,
                ),
                promoted_observation(
                    f"{row_id}_average",
                    project_id,
                    source_id,
                    item_code,
                    description,
                    unit_raw,
                    unit_normalized,
                    quantity,
                    average_unit_price,
                    "public_bid_tab_average",
                    date_basis,
                ),
            ]
        )

        bidder_unit_prices: list[Decimal] = []
        for bidder_group in bidder_groups:
            unit_price = decimal_value(sheet.cell(row_index, bidder_group.unit_col).value)
            extended_price = (quantity * unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            bid_id = f"{row_prefix}_{slugify(bidder_group.name)}"
            bidder_unit_prices.append(unit_price)
            bid_totals[bidder_group.name] += extended_price
            bidder_items.append(
                {
                    "bidder_item_observation_id": f"{row_id}_{slugify(bidder_group.name)}",
                    "bid_tab_item_id": row_id,
                    "bid_id": bid_id,
                    "project_id": project_id,
                    "source_id": source_id,
                    "agency_item_code": item_code,
                    "description_raw": description,
                    "unit_raw": unit_raw,
                    "unit_normalized": unit_normalized,
                    "quantity": decimal_text(quantity),
                    "unit_price": decimal_text(unit_price),
                    "extended_price": decimal_text(extended_price),
                }
            )

            validate_line_total(
                sheet.cell(row_index, bidder_group.total_col).value,
                formula_sheet.cell(row_index, bidder_group.total_col).value,
                extended_price,
                warnings,
                row_index,
                bidder_group.name,
            )

        calculated_average = (sum(bidder_unit_prices) / Decimal(len(bidder_unit_prices))).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        if abs(calculated_average - average_unit_price) > Decimal("0.01"):
            warnings.append(
                f"Row {row_index}: average bid {average_unit_price} differs from bidder average {calculated_average}"
            )

    ranked_bidders = sorted(bid_totals.items(), key=lambda item: item[1])
    bidder_bids = [
        {
            "bid_id": f"{row_prefix}_{slugify(name)}",
            "project_id": project_id,
            "source_id": source_id,
            "bidder_name": name,
            "bid_total": decimal_text(total),
            "bid_rank": str(index),
            "apparent_low": "true" if index == 1 else "false",
        }
        for index, (name, total) in enumerate(ranked_bidders, start=1)
    ]

    return ParsedBidTab(project_name, project_number, item_rows, bidder_bids, bidder_items, observations, warnings)


def parse_split_header_bid_tab_workbook(
    workbook_path: Path,
    sheet,
    formula_sheet,
    source_id: str,
    project_id: str,
    row_prefix: str,
    date_basis: str,
    group_row: int,
    header_row: int,
) -> ParsedBidTab:
    project_name = str(sheet["B1"].value or "").strip()
    project_number = split_header_project_number(str(sheet["B2"].value or "").strip())
    if not project_name or not project_number:
        raise ValueError("Split-header bid tab workbook must include project name in B1 and project number in B2.")

    groups = detect_split_header_price_groups(sheet, group_row, header_row)
    engineer_group = next((group for group in groups if group.role == "engineer"), None)
    average_group = next((group for group in groups if group.role == "average"), None)
    bidder_groups = [group for group in groups if group.role == "bidder"]

    if not engineer_group or not average_group:
        raise ValueError("Split-header bid tab workbook must include engineer estimate and average bid price groups.")
    if not bidder_groups:
        raise ValueError("Split-header bid tab workbook must include at least one bidder price group.")

    total_row = find_split_header_total_row(sheet, header_row + 1)
    item_rows: list[dict[str, str]] = []
    observations: list[dict[str, str]] = []
    bidder_items: list[dict[str, str]] = []
    bid_totals: dict[str, Decimal] = {group.name: Decimal("0.00") for group in bidder_groups}
    engineer_total = Decimal("0.00")
    warnings: list[str] = []

    for row_index in range(header_row + 1, total_row):
        item_code = normalize_bid_form_item_code(sheet.cell(row_index, 1).value)
        if not item_code:
            continue

        description = str(sheet.cell(row_index, 2).value or "").strip()
        unit_raw = str(sheet.cell(row_index, 3).value or "").strip()
        unit_normalized = normalize_unit(unit_raw)
        quantity = decimal_value(sheet.cell(row_index, 4).value)
        engineer_unit_price = decimal_value(sheet.cell(row_index, engineer_group.unit_col).value)
        engineer_extended_price = (quantity * engineer_unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        engineer_total += engineer_extended_price
        row_id = f"{row_prefix}_row_{row_index:04d}"

        bidder_unit_prices: list[Decimal] = []
        for bidder_group in bidder_groups:
            unit_price = decimal_value(sheet.cell(row_index, bidder_group.unit_col).value)
            extended_price = (quantity * unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            bid_id = f"{row_prefix}_{slugify(bidder_group.name)}"
            bidder_unit_prices.append(unit_price)
            bid_totals[bidder_group.name] += extended_price
            bidder_items.append(
                {
                    "bidder_item_observation_id": f"{row_id}_{slugify(bidder_group.name)}",
                    "bid_tab_item_id": row_id,
                    "bid_id": bid_id,
                    "project_id": project_id,
                    "source_id": source_id,
                    "agency_item_code": item_code,
                    "description_raw": description,
                    "unit_raw": unit_raw,
                    "unit_normalized": unit_normalized,
                    "quantity": decimal_text(quantity),
                    "unit_price": decimal_text(unit_price),
                    "extended_price": decimal_text(extended_price),
                }
            )
            validate_line_total(
                sheet.cell(row_index, bidder_group.total_col).value,
                formula_sheet.cell(row_index, bidder_group.total_col).value,
                extended_price,
                warnings,
                row_index,
                bidder_group.name,
            )

        validate_line_total(
            sheet.cell(row_index, engineer_group.total_col).value,
            formula_sheet.cell(row_index, engineer_group.total_col).value,
            engineer_extended_price,
            warnings,
            row_index,
            "Engineer estimate",
        )

        average_unit_price = (sum(bidder_unit_prices) / Decimal(len(bidder_unit_prices))).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        workbook_average = decimal_value(sheet.cell(row_index, average_group.unit_col).value)
        if abs(average_unit_price - workbook_average) > Decimal("0.01"):
            warnings.append(
                f"Row {row_index}: average bid {workbook_average} differs from bidder average {average_unit_price}"
            )

        item_rows.append(
            bid_tab_item_row(
                row_id,
                project_id,
                source_id,
                workbook_path,
                sheet.title,
                row_index,
                project_number,
                "",
                item_code,
                "CDOT",
                item_code,
                description,
                item_code,
                unit_raw,
                unit_normalized,
                quantity,
                engineer_unit_price,
                average_unit_price,
                date_basis,
            )
        )
        observations.extend(
            [
                promoted_observation(
                    f"{row_id}_engineer_estimate",
                    project_id,
                    source_id,
                    item_code,
                    description,
                    unit_raw,
                    unit_normalized,
                    quantity,
                    engineer_unit_price,
                    "public_bid_tab_engineer_estimate",
                    date_basis,
                ),
                promoted_observation(
                    f"{row_id}_average",
                    project_id,
                    source_id,
                    item_code,
                    description,
                    unit_raw,
                    unit_normalized,
                    quantity,
                    average_unit_price,
                    "public_bid_tab_average",
                    date_basis,
                ),
            ]
        )

    validate_published_total(sheet, total_row, engineer_group, engineer_total, warnings)
    for bidder_group in bidder_groups:
        validate_published_total(sheet, total_row, bidder_group, bid_totals[bidder_group.name], warnings)

    ranked_bidders = sorted(bid_totals.items(), key=lambda item: item[1])
    bidder_bids = [
        {
            "bid_id": f"{row_prefix}_{slugify(name)}",
            "project_id": project_id,
            "source_id": source_id,
            "bidder_name": name,
            "bid_total": decimal_text(total),
            "bid_rank": str(index),
            "apparent_low": "true" if index == 1 else "false",
        }
        for index, (name, total) in enumerate(ranked_bidders, start=1)
    ]

    return ParsedBidTab(project_name, project_number, item_rows, bidder_bids, bidder_items, observations, warnings)


def parse_bid_form_workbook(
    workbook_path: Path,
    workbook,
    formula_workbook,
    source_id: str,
    project_id: str,
    row_prefix: str,
    date_basis: str,
    header_row: int,
) -> ParsedBidTab:
    sheet = workbook["Bid Form"] if "Bid Form" in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    formula_sheet = formula_workbook[sheet.title]
    column_by_header = bid_form_column_map(sheet, header_row)
    groups = detect_bid_form_price_groups(sheet, header_row)
    engineer_group = next((group for group in groups if group.role == "engineer"), None)
    bidder_groups = [group for group in groups if group.role == "bidder"]

    if not engineer_group:
        raise ValueError("Bid-form workbook must include estimated unit price and total cost columns.")
    if not bidder_groups:
        raise ValueError("Bid-form workbook must include at least one bidder unit price / total cost column pair.")

    project_number, project_name, summary_engineer_total = bid_form_summary_metadata(workbook)
    if not project_number or not project_name:
        raise ValueError("Bid-form workbook must include project metadata on Sheet1.")

    item_rows: list[dict[str, str]] = []
    observations: list[dict[str, str]] = []
    bidder_items: list[dict[str, str]] = []
    bid_totals: dict[str, Decimal] = {group.name: Decimal("0.00") for group in bidder_groups}
    engineer_itemized_total = Decimal("0.00")
    warnings: list[str] = []

    for row_index in range(header_row + 1, sheet.max_row + 1):
        item_code_value = sheet.cell(row_index, column_by_header["ITEM NO."]).value
        item_code = normalize_bid_form_item_code(item_code_value)
        if not item_code:
            continue

        description = str(sheet.cell(row_index, column_by_header["ITEM DESCRIPTION"]).value or "").strip()
        quantity = decimal_value(sheet.cell(row_index, column_by_header["QUANTITY"]).value)
        unit_raw = str(sheet.cell(row_index, column_by_header["UNIT"]).value or "").strip()
        unit_normalized = normalize_unit(unit_raw)
        engineer_unit_price = decimal_value(sheet.cell(row_index, engineer_group.unit_col).value)
        engineer_extended_price = (quantity * engineer_unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        engineer_itemized_total += engineer_extended_price

        bidder_unit_prices: list[Decimal] = []
        row_id = f"{row_prefix}_row_{row_index:04d}"

        for bidder_group in bidder_groups:
            unit_price = decimal_value(sheet.cell(row_index, bidder_group.unit_col).value)
            extended_price = (quantity * unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            bid_id = f"{row_prefix}_{slugify(bidder_group.name)}"
            bidder_unit_prices.append(unit_price)
            bid_totals[bidder_group.name] += extended_price
            bidder_items.append(
                {
                    "bidder_item_observation_id": f"{row_id}_{slugify(bidder_group.name)}",
                    "bid_tab_item_id": row_id,
                    "bid_id": bid_id,
                    "project_id": project_id,
                    "source_id": source_id,
                    "agency_item_code": item_code,
                    "description_raw": description,
                    "unit_raw": unit_raw,
                    "unit_normalized": unit_normalized,
                    "quantity": decimal_text(quantity),
                    "unit_price": decimal_text(unit_price),
                    "extended_price": decimal_text(extended_price),
                }
            )

            validate_line_total(
                sheet.cell(row_index, bidder_group.total_col).value,
                formula_sheet.cell(row_index, bidder_group.total_col).value,
                extended_price,
                warnings,
                row_index,
                bidder_group.name,
            )

        validate_line_total(
            sheet.cell(row_index, engineer_group.total_col).value,
            formula_sheet.cell(row_index, engineer_group.total_col).value,
            engineer_extended_price,
            warnings,
            row_index,
            "Engineer estimate",
        )

        average_unit_price = (sum(bidder_unit_prices) / Decimal(len(bidder_unit_prices))).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )
        item_rows.append(
            bid_tab_item_row(
                row_id,
                project_id,
                source_id,
                workbook_path,
                sheet.title,
                row_index,
                project_number,
                "",
                item_code,
                source_code_system(item_code),
                item_code,
                description,
                item_code,
                unit_raw,
                unit_normalized,
                quantity,
                engineer_unit_price,
                average_unit_price,
                date_basis,
            )
        )

        observations.extend(
            [
                promoted_observation(
                    f"{row_id}_engineer_estimate",
                    project_id,
                    source_id,
                    item_code,
                    description,
                    unit_raw,
                    unit_normalized,
                    quantity,
                    engineer_unit_price,
                    "public_bid_tab_engineer_estimate",
                    date_basis,
                ),
                promoted_observation(
                    f"{row_id}_average",
                    project_id,
                    source_id,
                    item_code,
                    description,
                    unit_raw,
                    unit_normalized,
                    quantity,
                    average_unit_price,
                    "public_bid_tab_average",
                    date_basis,
                ),
            ]
        )

    if summary_engineer_total is not None and abs(summary_engineer_total - engineer_itemized_total) > Decimal("0.01"):
        warnings.append(
            "Sheet1 engineer estimate "
            f"{summary_engineer_total} differs from itemized engineer estimate total {engineer_itemized_total}"
        )

    ranked_bidders = sorted(bid_totals.items(), key=lambda item: item[1])
    bidder_bids = [
        {
            "bid_id": f"{row_prefix}_{slugify(name)}",
            "project_id": project_id,
            "source_id": source_id,
            "bidder_name": name,
            "bid_total": decimal_text(total),
            "bid_rank": str(index),
            "apparent_low": "true" if index == 1 else "false",
        }
        for index, (name, total) in enumerate(ranked_bidders, start=1)
    ]

    return ParsedBidTab(project_name, project_number, item_rows, bidder_bids, bidder_items, observations, warnings)


def find_header_row(sheet) -> int:
    for row_index in range(1, sheet.max_row + 1):
        if str(sheet.cell(row_index, 3).value or "").strip().upper() == "ITEM CODE":
            return row_index
    raise ValueError("Could not find ITEM CODE header in column C.")


def find_results_header_row(sheet) -> int | None:
    required_headers = {"NO.", "SPEC*", "ITEM", "UNIT", "QUANTITY"}
    for row_index in range(1, min(sheet.max_row, 25) + 1):
        headers = {str(sheet.cell(row_index, column).value or "").strip().upper() for column in range(1, sheet.max_column + 1)}
        if required_headers.issubset(headers):
            return row_index
    return None


def results_column_map(sheet, header_row: int) -> dict[str, int]:
    columns = {
        str(sheet.cell(header_row, column).value or "").strip().upper(): column
        for column in range(1, sheet.max_column + 1)
    }
    required_headers = ["NO.", "SPEC*", "ITEM", "UNIT", "QUANTITY"]
    missing = [header for header in required_headers if header not in columns]
    if missing:
        raise ValueError(f"Results workbook missing required header(s): {', '.join(missing)}")
    return columns


def results_reconciliation(sheet, column_by_header: dict[str, int], row_index: int) -> dict[str, str]:
    raw_item_code = cell_text(sheet, row_index, column_by_header.get("CDOT ITEM CODE"))
    item_code = "" if raw_item_code.lower() == "none" else raw_item_code.upper()
    description = cell_text(sheet, row_index, column_by_header.get("CDOT DESCRIPTION"))
    unit_raw = cell_text(sheet, row_index, column_by_header.get("CDOT UNIT"))

    if item_code and not ITEM_CODE_PATTERN.match(item_code):
        raise ValueError(f"Row {row_index}: CDOT Item Code {item_code!r} is not shaped like ###-#####.")

    return {
        "item_code": item_code,
        "description": "" if description.lower() == "none" else description,
        "unit_raw": "" if unit_raw.lower() == "none" else unit_raw,
    }


def cell_text(sheet, row_index: int, column_index: int | None) -> str:
    if not column_index:
        return ""
    return str(sheet.cell(row_index, column_index).value or "").strip()


def find_results_total_row(sheet, item_column: int, start_row: int) -> int:
    for row_index in range(start_row, sheet.max_row + 1):
        if str(sheet.cell(row_index, item_column).value or "").strip().upper() == "TOTAL":
            return row_index
    raise ValueError("Results workbook must include a TOTAL row in the Item column.")


def find_bid_form_header_row(sheet) -> int | None:
    required_headers = {"ITEM NO.", "ITEM DESCRIPTION", "UNIT", "QUANTITY"}
    for row_index in range(1, min(sheet.max_row, 25) + 1):
        headers = {str(sheet.cell(row_index, column).value or "").strip().upper() for column in range(1, sheet.max_column + 1)}
        if required_headers.issubset(headers):
            return row_index
    return None


def find_split_bid_tab_header_rows(sheet) -> tuple[int, int] | None:
    if str(sheet.cell(4, 2).value or "").strip().upper() != "TABULATION OF BIDS":
        return None

    for row_index in range(1, min(sheet.max_row, 25)):
        if (
            str(sheet.cell(row_index, 1).value or "").strip().upper() == "ITEM"
            and str(sheet.cell(row_index + 1, 1).value or "").strip().upper() == "NO."
            and str(sheet.cell(row_index + 1, 2).value or "").strip().upper() == "ITEM"
            and str(sheet.cell(row_index + 1, 3).value or "").strip().upper() == "UNIT"
            and str(sheet.cell(row_index + 1, 4).value or "").strip().upper() == "QUANTITY"
        ):
            return row_index - 2, row_index + 1
    return None


def split_header_project_number(value: str) -> str:
    match = re.search(r"PROJECT\s+NO\.\s*(?P<value>.+)$", value, re.IGNORECASE)
    return match.group("value").strip() if match else value.strip()


def detect_split_header_price_groups(sheet, group_row: int, header_row: int) -> list[PriceColumnGroup]:
    groups: list[PriceColumnGroup] = []
    for column in range(1, sheet.max_column):
        unit_header = str(sheet.cell(header_row - 1, column).value or "").strip().upper()
        total_header = str(sheet.cell(header_row - 1, column + 1).value or "").strip().upper()
        unit_subheader = str(sheet.cell(header_row, column).value or "").strip().upper()
        total_subheader = str(sheet.cell(header_row, column + 1).value or "").strip().upper()
        if unit_header != "UNIT" or total_header != "EXTENDED" or unit_subheader != "PRICE" or total_subheader != "COST":
            continue

        name = str(sheet.cell(group_row, column).value or "").strip()
        if not name:
            continue

        upper_name = name.upper()
        if "ENGINEER" in upper_name:
            role = "engineer"
        elif "AVERAGE" in upper_name:
            role = "average"
        else:
            role = "bidder"
        groups.append(PriceColumnGroup("Engineer Estimate" if role == "engineer" else name, column, column + 1, role))
    return groups


def find_split_header_total_row(sheet, start_row: int) -> int:
    for row_index in range(start_row, sheet.max_row + 1):
        if str(sheet.cell(row_index, 2).value or "").strip().upper() == "BID SCHEDULE TOTAL":
            return row_index
    raise ValueError("Split-header bid tab workbook must include a BID SCHEDULE TOTAL row.")


def validate_published_total(
    sheet,
    total_row: int,
    group: PriceColumnGroup,
    calculated_total: Decimal,
    warnings: list[str],
) -> None:
    published_total = decimal_value(sheet.cell(total_row, group.total_col).value)
    if abs(published_total - calculated_total) > Decimal("0.01"):
        warnings.append(
            f"Bid schedule total {group.name} {published_total} differs from calculated total {calculated_total}"
        )


def bid_form_column_map(sheet, header_row: int) -> dict[str, int]:
    columns = {
        str(sheet.cell(header_row, column).value or "").strip().upper(): column
        for column in range(1, sheet.max_column + 1)
    }
    required_headers = ["ITEM NO.", "ITEM DESCRIPTION", "UNIT", "QUANTITY"]
    missing = [header for header in required_headers if header not in columns]
    if missing:
        raise ValueError(f"Bid-form workbook missing required header(s): {', '.join(missing)}")
    return columns


def detect_price_groups(sheet, group_row: int, header_row: int) -> list[PriceColumnGroup]:
    groups: list[PriceColumnGroup] = []
    for column in range(1, sheet.max_column):
        title = sheet.cell(group_row, column).value
        unit_header = str(sheet.cell(header_row, column).value or "").strip().upper()
        total_header = str(sheet.cell(header_row, column + 1).value or "").strip().upper()
        if not title or unit_header != "UNIT COST" or total_header != "TOTAL COST":
            continue

        name = str(title).strip()
        upper_name = name.upper()
        if "ENGINEER" in upper_name:
            role = "engineer"
        elif "AVERAGE" in upper_name:
            role = "average"
        else:
            role = "bidder"
        groups.append(PriceColumnGroup(name, column, column + 1, role))
    return groups


def detect_results_price_groups(sheet, group_row: int, header_row: int) -> list[PriceColumnGroup]:
    groups: list[PriceColumnGroup] = []
    for column in range(1, sheet.max_column):
        unit_header = str(sheet.cell(header_row, column).value or "").strip().upper()
        total_header = str(sheet.cell(header_row, column + 1).value or "").strip().upper()
        if unit_header != "UNIT COST" or total_header != "EXTENDED COST":
            continue

        name = str(sheet.cell(group_row, column).value or "").strip()
        if not name:
            continue

        role = "engineer" if "ENGINEER" in name.upper() else "bidder"
        groups.append(PriceColumnGroup(name, column, column + 1, role))
    return groups


def detect_bid_form_price_groups(sheet, header_row: int) -> list[PriceColumnGroup]:
    groups: list[PriceColumnGroup] = []
    for column in range(1, sheet.max_column):
        unit_header = str(sheet.cell(header_row, column).value or "").strip()
        total_header = str(sheet.cell(header_row, column + 1).value or "").strip().upper()
        if not unit_header.upper().endswith("UNIT PRICE") or total_header != "TOTAL COST BASE BID":
            continue

        name = re.sub(r"\s*UNIT PRICE\s*$", "", unit_header, flags=re.IGNORECASE).strip()
        role = "engineer" if name.upper() == "ESTIMATED" else "bidder"
        groups.append(PriceColumnGroup("Engineer Estimate" if role == "engineer" else name, column, column + 1, role))
    return groups


def normalize_bid_form_item_code(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)) and Decimal(str(value)) == Decimal(str(value)).to_integral_value():
        text = str(int(value))
    else:
        text = str(value).strip()
    if BID_FORM_ITEM_CODE_PATTERN.match(text):
        return text.upper()
    return ""


def normalize_source_item_number(value: object) -> str:
    if isinstance(value, (int, float)) and Decimal(str(value)) == Decimal(str(value)).to_integral_value():
        return str(int(value))
    return str(value).strip()


def bid_form_summary_metadata(workbook) -> tuple[str, str, Decimal | None]:
    project_number = ""
    project_name = ""
    engineer_total: Decimal | None = None

    if "Sheet1" not in workbook.sheetnames:
        return project_number, project_name, engineer_total

    sheet = workbook["Sheet1"]
    project_number = str(sheet["A1"].value or "").strip()
    project_name = str(sheet["A2"].value or "").strip()
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            if str(cell.value or "").strip().upper() == "ENGR EST":
                for candidate_column in range(cell.column + 1, sheet.max_column + 1):
                    value = sheet.cell(cell.row, candidate_column).value
                    if value is not None:
                        engineer_total = decimal_value(value)
                        break
                break

    return project_number, project_name, engineer_total


def validate_line_total(
    cached_total: object,
    formula_or_value: object,
    calculated_total: Decimal,
    warnings: list[str],
    row_index: int,
    party_name: str,
) -> None:
    if isinstance(formula_or_value, str) and formula_or_value.startswith("="):
        return
    if cached_total is None:
        return
    parsed_total = decimal_value(cached_total)
    if abs(parsed_total - calculated_total) > Decimal("0.01"):
        warnings.append(f"Row {row_index}: {party_name} total {parsed_total} differs from calculated {calculated_total}")


def promoted_observation(
    observation_id: str,
    project_id: str,
    source_id: str,
    item_code: str,
    description: str,
    unit_raw: str,
    unit_normalized: str,
    quantity: Decimal,
    unit_price: Decimal,
    price_type: str,
    date_basis: str,
) -> dict[str, str]:
    return {
        "observation_id": observation_id,
        "project_id": project_id,
        "source_id": source_id,
        "agency_item_code": item_code,
        "description_raw": description,
        "description_normalized": normalize_description(description),
        "unit_raw": unit_raw,
        "unit_normalized": unit_normalized,
        "quantity": quantity_text(quantity),
        "unit_price": decimal_text(unit_price),
        "extended_price": decimal_text(quantity * unit_price),
        "discipline": "Roadway",
        "price_type": price_type,
        "date_basis": date_basis,
    }


def optional_decimal(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        cleaned = value.replace("$", "").replace(",", "").strip()
        match = re.search(r"-?\d+(?:\.\d+)?", cleaned)
        if not match:
            return None
        value = match.group(0)
    try:
        return decimal_value(value)
    except ValueError:
        return None


def optional_quantity(value: object) -> Decimal | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        cleaned = value.replace(",", "").strip()
        if not re.fullmatch(r"-?\d+(?:\.\d+)?", cleaned):
            return None
        value = cleaned
    try:
        return Decimal(str(value))
    except InvalidOperation:
        return None


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def configured_rows(config: dict[str, Any]) -> Iterable[int]:
    for first, last in config["row_ranges"]:
        yield from range(int(first), int(last) + 1)


def configured_total(sheet, bidder: dict[str, Any]) -> Decimal | None:
    total_cell = bidder.get("published_total_cell")
    return optional_decimal(sheet[total_cell].value) if total_cell else None


def parse_master_sheet(
    workbook_path: Path,
    sheet,
    config: dict[str, Any],
) -> ParsedBidTab:
    source_id = config["source_id"]
    project_id = config["project_id"]
    row_prefix = config["row_prefix"]
    date_basis = config["source_date"]
    columns = config["columns"]
    layout = config["layout"]
    bidders = config.get("bidders", [])
    item_rows: list[dict[str, str]] = []
    bidder_items: list[dict[str, str]] = []
    observations: list[dict[str, str]] = []
    warnings: list[str] = []
    calculated_bid_totals = {bidder["name"]: Decimal("0.00") for bidder in bidders}

    for row_index in configured_rows(config):
        raw_code = sheet.cell(row_index, int(columns["code"])).value
        description = re.sub(r"\s+", " ", str(sheet.cell(row_index, int(columns["description"])).value or "")).strip()
        if raw_code is None or not description:
            continue
        source_item_code = normalize_source_item_number(raw_code).upper()
        if not source_item_code or not source_item_code[0].isdigit():
            continue

        unit_raw = str(sheet.cell(row_index, int(columns["unit"])).value or "").strip()
        unit_normalized = normalize_unit(unit_raw)
        raw_quantity = sheet.cell(row_index, int(columns["quantity"])).value
        quantity = optional_quantity(raw_quantity)
        if quantity is None:
            quantity = Decimal("1.00")
            warnings.append(f"Row {row_index}: nonnumeric quantity {raw_quantity!r} treated as 1 for source total reconciliation.")

        code_system = source_code_system(source_item_code)
        matched_code = source_item_code if ITEM_CODE_PATTERN.match(source_item_code) else ""
        match_status = "matched" if matched_code else "source_cdot_prefix_only" if re.fullmatch(r"\d{3}", source_item_code) else "unmatched"
        row_id = f"{row_prefix}_row_{row_index:04d}"
        engineer_quantity = quantity
        engineer_unit_price: Decimal | None = None
        engineer_total: Decimal | None = None
        if columns.get("engineer_quantity"):
            engineer_quantity = optional_quantity(sheet.cell(row_index, int(columns["engineer_quantity"])).value) or quantity
        if columns.get("engineer_unit"):
            engineer_unit_price = optional_decimal(sheet.cell(row_index, int(columns["engineer_unit"])).value)
        if columns.get("engineer_total"):
            engineer_total = optional_decimal(sheet.cell(row_index, int(columns["engineer_total"])).value)
        if engineer_unit_price is None and engineer_total is not None and engineer_quantity != 0:
            engineer_unit_price = (engineer_total / engineer_quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

        bidder_unit_prices: list[Decimal] = []
        for bidder in bidders:
            bidder_quantity = optional_quantity(sheet.cell(row_index, int(bidder.get("quantity_col", columns["quantity"]))).value) or quantity
            unit_price = optional_decimal(sheet.cell(row_index, int(bidder["unit_col"])).value)
            published_line_total = optional_decimal(sheet.cell(row_index, int(bidder["total_col"])).value)
            if unit_price is None and published_line_total is not None and bidder_quantity != 0:
                unit_price = (published_line_total / bidder_quantity).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if unit_price is None:
                continue
            calculated_line_total = (bidder_quantity * unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            bidder_unit_prices.append(unit_price)
            calculated_bid_totals[bidder["name"]] += calculated_line_total
            bid_id = f"{row_prefix}_{slugify(bidder['name'])}"
            bidder_items.append({
                "bidder_item_observation_id": f"{row_id}_{slugify(bidder['name'])}",
                "bid_tab_item_id": row_id,
                "bid_id": bid_id,
                "project_id": project_id,
                "source_id": source_id,
                "agency_item_code": matched_code,
                "description_raw": description,
                "unit_raw": unit_raw,
                "unit_normalized": unit_normalized,
                "quantity": quantity_text(bidder_quantity),
                "unit_price": decimal_text(unit_price),
                "extended_price": decimal_text(calculated_line_total),
            })
            if published_line_total is not None and abs(published_line_total - calculated_line_total) > Decimal("0.01"):
                warnings.append(
                    f"Row {row_index}: {bidder['name']} source line total {published_line_total} "
                    f"differs from calculated {calculated_line_total}."
                )

        average_unit_price = (
            (sum(bidder_unit_prices) / Decimal(len(bidder_unit_prices))).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if bidder_unit_prices else None
        )
        item_rows.append(bid_tab_item_row(
            row_id, project_id, source_id, workbook_path, sheet.title, row_index, config["project_number"],
            str(len(item_rows) + 1), source_item_code, code_system, source_item_code, description,
            matched_code or source_item_code, unit_raw, unit_normalized, quantity,
            engineer_unit_price or Decimal("0.00"), average_unit_price or Decimal("0.00"), date_basis,
            matched_code, match_status,
        ))
        item_rows[-1]["engineer_estimate_unit_price"] = decimal_text(engineer_unit_price) if engineer_unit_price is not None else ""
        item_rows[-1]["average_bid_unit_price"] = decimal_text(average_unit_price) if average_unit_price is not None else ""
        item_rows[-1]["quantity"] = quantity_text(quantity)
        item_rows[-1]["engineer_estimate_quantity"] = quantity_text(engineer_quantity) if engineer_unit_price is not None else ""
        audit_unit_price = optional_decimal(
            sheet.cell(row_index, int(columns["audit_engineer_unit"])).value
        ) if columns.get("audit_engineer_unit") else None
        audit_total = optional_decimal(
            sheet.cell(row_index, int(columns["audit_engineer_total"])).value
        ) if columns.get("audit_engineer_total") else None
        item_rows[-1]["audit_estimate_unit_price"] = decimal_text(audit_unit_price) if audit_unit_price is not None else ""
        item_rows[-1]["audit_estimate_extended_price"] = decimal_text(audit_total) if audit_total is not None else ""

        if matched_code and engineer_unit_price is not None:
            observations.append(promoted_observation(
                f"{row_id}_engineer_estimate", project_id, source_id, matched_code, description, unit_raw,
                unit_normalized, engineer_quantity, engineer_unit_price,
                "engineers_estimate" if layout == "estimate" else "public_bid_tab_engineer_estimate", date_basis,
            ))
        if matched_code and average_unit_price is not None:
            observations.append(promoted_observation(
                f"{row_id}_average", project_id, source_id, matched_code, description, unit_raw,
                unit_normalized, quantity, average_unit_price, "public_bid_tab_average", date_basis,
            ))
        if engineer_total is not None and engineer_unit_price is not None:
            calculated_engineer_total = (engineer_quantity * engineer_unit_price).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
            if abs(engineer_total - calculated_engineer_total) > Decimal("0.01"):
                warnings.append(
                    f"Row {row_index}: engineer source line total {engineer_total} differs from calculated {calculated_engineer_total}."
                )

    ranked_totals: list[tuple[str, Decimal]] = []
    for bidder in bidders:
        published_total = configured_total(sheet, bidder)
        calculated_total = calculated_bid_totals[bidder["name"]]
        ranking_total = published_total if published_total is not None else calculated_total
        ranked_totals.append((bidder["name"], ranking_total))
        if published_total is not None and abs(published_total - calculated_total) > Decimal("0.01"):
            warnings.append(
                f"Included schedule total {bidder['name']} {published_total} differs from calculated total {calculated_total}."
            )

    awarded_bidder = config.get("awarded_bidder", "")
    bidder_bids = [{
        "bid_id": f"{row_prefix}_{slugify(name)}",
        "project_id": project_id,
        "source_id": source_id,
        "bidder_name": name,
        "bid_total": decimal_text(total),
        "bid_rank": str(rank),
        "apparent_low": "true" if rank == 1 else "false",
        "is_awarded": "true" if name == awarded_bidder else "false",
    } for rank, (name, total) in enumerate(sorted(ranked_totals, key=lambda pair: pair[1]), start=1)]

    expected_count = int(config["expected_item_count"])
    if len(item_rows) != expected_count:
        raise ValueError(f"{sheet.title}: expected {expected_count} source items, parsed {len(item_rows)}.")
    expected_bidders = int(config.get("expected_bidder_count", len(bidders)))
    if len(bidder_bids) != expected_bidders:
        raise ValueError(f"{sheet.title}: expected {expected_bidders} bidders, parsed {len(bidder_bids)}.")
    return ParsedBidTab(
        config["project_name"], config["project_number"], item_rows, bidder_bids, bidder_items, observations, warnings
    )


def add_resolved_master_observations(parsed: ParsedBidTab, source_type_value: str) -> None:
    existing = {row["observation_id"] for row in parsed.observations}
    bidder_items_by_row: dict[str, list[dict[str, str]]] = {}
    for bidder_item in parsed.bidder_items:
        bidder_items_by_row.setdefault(bidder_item["bid_tab_item_id"], []).append(bidder_item)
    awarded_bid_ids = {bid["bid_id"] for bid in parsed.bidder_bids if bid.get("is_awarded") == "true"}
    parsed.observations[:] = [
        row for row in parsed.observations
        if any(item["matched_agency_item_code"] == row["agency_item_code"] for item in parsed.item_rows)
    ]
    for item in parsed.item_rows:
        code = item["matched_agency_item_code"]
        if item["match_status"] != "matched" or not code:
            continue
        engineer_price = optional_decimal(item["engineer_estimate_unit_price"])
        engineer_id = f"{item['bid_tab_item_id']}_engineer_estimate"
        if engineer_price is not None and engineer_id not in existing:
            parsed.observations.append(promoted_observation(
                engineer_id, item["project_id"], item["source_id"], code, item["source_item_description"],
                item["unit_raw"], item["unit_normalized"],
                optional_quantity(item.get("engineer_estimate_quantity")) or optional_quantity(item["quantity"]) or Decimal("0"),
                engineer_price,
                "engineers_estimate" if source_type_value == "estimate" else "public_bid_tab_engineer_estimate",
                item["date_basis"],
            ))
        average_price = optional_decimal(item["average_bid_unit_price"])
        average_id = f"{item['bid_tab_item_id']}_average"
        if average_price is not None and average_id not in existing:
            parsed.observations.append(promoted_observation(
                average_id, item["project_id"], item["source_id"], code, item["source_item_description"],
                item["unit_raw"], item["unit_normalized"], decimal_value(item["quantity"]), average_price,
                "public_bid_tab_average", item["date_basis"],
            ))
        awarded_item = next(
            (candidate for candidate in bidder_items_by_row.get(item["bid_tab_item_id"], []) if candidate["bid_id"] in awarded_bid_ids),
            None,
        )
        awarded_id = f"{item['bid_tab_item_id']}_awarded"
        if awarded_item and awarded_id not in existing:
            parsed.observations.append(promoted_observation(
                awarded_id, item["project_id"], item["source_id"], code, item["source_item_description"],
                awarded_item["unit_raw"], awarded_item["unit_normalized"],
                optional_quantity(awarded_item["quantity"]) or Decimal("0"), decimal_value(awarded_item["unit_price"]),
                "public_bid_tab_awarded", item["date_basis"],
            ))


def source_row(args: argparse.Namespace) -> dict[str, str]:
    return {
        "source_id": args.source_id,
        "source_type": args.source_type,
        "agency": args.agency_owner,
        "state": args.state,
        "source_label": args.source_label,
        "source_date": args.date_basis,
        "data_year": args.source_year,
        "source_url": getattr(args, "source_url", ""),
        "source_file_name": Path(args.workbook).name,
        "sha256": file_sha256(Path(args.workbook)),
        "parser_name": "import_bid_tab_workbook.py",
        "parser_version": "3.0.0",
        "notes": getattr(
            args,
            "source_notes",
            f"Public bid tab workbook promoted from {Path(args.workbook).name}. Apparent low is not confirmed award.",
        ),
    }


def project_row(args: argparse.Namespace, parsed: ParsedBidTab) -> dict[str, str]:
    awarded_bid = next((bid for bid in parsed.bidder_bids if bid.get("is_awarded") == "true"), None)
    return {
        "project_id": args.project_id,
        "project_name": parsed.project_name,
        "agency_owner": args.agency_owner,
        "state": args.state,
        "county_region": args.county_region,
        "work_type": args.work_type,
        "estimate_let_date": args.date_basis,
        "source_id": args.source_id,
        "project_number": parsed.project_number,
        "project_location_raw": parsed.project_name,
        "contractor": awarded_bid["bidder_name"] if awarded_bid else "",
        "district": args.district,
        "terrain": "",
        "bid_count": str(len(parsed.bidder_bids)),
        "awarded_bid_total": awarded_bid["bid_total"] if awarded_bid else "",
        "award_index": "",
    }


def validate_import(
    parsed: ParsedBidTab,
    known_item_codes: set[str],
    allow_missing_item_codes: bool = False,
    source_only: bool = False,
) -> tuple[list[str], list[str]]:
    errors: list[str] = []
    warnings = list(parsed.warnings)

    if not parsed.item_rows:
        errors.append("No item rows were parsed.")
    if not parsed.bidder_bids and any(row.get("average_bid_unit_price") for row in parsed.item_rows):
        errors.append("No bidder columns were parsed.")

    if source_only:
        invalid_statuses = sorted(
            {
                row.get("match_status", "")
                for row in parsed.item_rows
                if row.get("match_status", "") not in {"matched", "unmatched", "source_cdot_prefix_only"}
            }
        )
        if invalid_statuses:
            errors.append(f"Invalid match status value(s): {', '.join(invalid_statuses)}")
        return errors, warnings

    missing_codes = sorted(
        {
            row["item_code"]
            for row in parsed.item_rows
            if row.get("match_status", "matched") == "matched"
            and row.get("item_code", "")
            and row["item_code"] not in known_item_codes
        }
    )
    if missing_codes:
        message = f"Missing agency item codes: {', '.join(missing_codes)}"
        if allow_missing_item_codes:
            warnings.append(message)
        else:
            errors.append(message)

    unknown_units = sorted(
        {
            row["unit_raw"]
            for row in parsed.item_rows
            if row["unit_raw"].strip().upper() == row["unit_normalized"] and row["unit_normalized"] not in known_normal_units()
        }
    )
    if unknown_units:
        warnings.append(f"Unknown normalized units preserved as raw values: {', '.join(unknown_units)}")

    return errors, warnings


def build_match_candidates(parsed: ParsedBidTab, agency_items: list[dict[str, str]]) -> list[dict[str, str]]:
    candidate_map: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in agency_items:
        unit = normalize_unit(row.get("official_unit", ""))
        for description_field in ["official_description", "official_abbreviated_description"]:
            description = row.get(description_field, "")
            if not description:
                continue
            candidate_map.setdefault((normalize_description(description), unit), []).append(row)

    candidates: list[dict[str, str]] = []
    for row in parsed.item_rows:
        key = (normalize_description(row.get("source_item_description") or row["item_description"]), row["unit_normalized"])
        matches = candidate_map.get(key, [])
        unique_matches = {
            match["item_code"].strip().upper(): match
            for match in matches
            if match.get("item_code")
        }
        if not unique_matches:
            continue

        status = "single_exact_description_unit_candidate" if len(unique_matches) == 1 else "multiple_exact_description_unit_candidates"
        for item_code, match in sorted(unique_matches.items()):
            candidates.append(
                {
                    "bid_tab_item_id": row["bid_tab_item_id"],
                    "source_item_code": row["source_item_code"],
                    "source_item_description": row.get("source_item_description") or row["item_description"],
                    "unit_normalized": row["unit_normalized"],
                    "candidate_agency_item_code": item_code,
                    "candidate_description": match.get("official_description", ""),
                    "candidate_unit": normalize_unit(match.get("official_unit", "")),
                    "candidate_count": str(len(unique_matches)),
                    "suggestion_status": status,
                }
            )
    return candidates


def resolve_short_item_codes(parsed: ParsedBidTab, agency_items: list[dict[str, str]]) -> list[str]:
    warnings: list[str] = []
    candidate_map: dict[tuple[str, str], list[dict[str, str]]] = {}
    for row in agency_items:
        unit = normalize_unit(row.get("official_unit", ""))
        for description_field in ["official_description", "official_abbreviated_description"]:
            description = row.get(description_field, "")
            if not description:
                continue
            candidate_map.setdefault((normalize_description(description), unit), []).append(row)

    replacements: dict[tuple[str, str, str], str] = {}
    for row in parsed.item_rows:
        item_code = row["item_code"]
        if not item_code or row.get("match_status") == "unmatched":
            continue
        if "-" in item_code:
            continue

        key = (normalize_description(row["item_description"]), row["unit_normalized"])
        matches = {
            candidate["item_code"].strip().upper()
            for candidate in candidate_map.get(key, [])
            if candidate.get("item_code", "").strip().upper().startswith(f"{item_code}-")
        }
        replacement_key = (item_code, row["item_description"], row["unit_normalized"])
        if len(matches) == 1:
            replacement = next(iter(matches))
            replacements[replacement_key] = replacement
            warnings.append(
                f"Resolved short agency item code {item_code} for {row['item_description']} ({row['unit_normalized']}) to {replacement}."
            )
        else:
            warnings.append(
                f"Preserved nonstandard item code {item_code} for {row['item_description']} ({row['unit_normalized']}); no exact compatible CDOT item match."
            )

    if not replacements:
        return warnings

    for row in parsed.item_rows:
        replacement = replacements.get((row["item_code"], row["item_description"], row["unit_normalized"]))
        if replacement:
            row["item_code"] = replacement
            row["matched_agency_item_code"] = replacement
            row["match_status"] = "matched"
    for row in parsed.observations:
        replacement = replacements.get((row["agency_item_code"], row["description_raw"], row["unit_normalized"]))
        if replacement:
            row["agency_item_code"] = replacement
    for row in parsed.bidder_items:
        replacement = replacements.get((row["agency_item_code"], row["description_raw"], row["unit_normalized"]))
        if replacement:
            row["agency_item_code"] = replacement

    return warnings


def known_normal_units() -> set[str]:
    return {"EACH", "L S", "LB", "SF", "SY", "CY", "LF", "ACRE", "HOUR", "DAY", "GAL", "CF", "TON", "F A"}


def promote(args: argparse.Namespace, parsed_override: ParsedBidTab | None = None) -> None:
    parsed = parsed_override or parse_workbook(Path(args.workbook), args.source_id, args.project_id, args.row_prefix, args.date_basis)
    item_rows = agency_item_rows(args.agency_items)
    mapping_warnings = [] if args.source_only else resolve_short_item_codes(parsed, item_rows)
    if parsed_override is not None and not args.source_only:
        add_resolved_master_observations(parsed, args.source_type)
    match_candidates = build_match_candidates(parsed, item_rows)
    errors, warnings = validate_import(
        parsed,
        {row["item_code"].strip().upper() for row in item_rows if row.get("item_code")},
        args.allow_missing_item_codes,
        args.source_only,
    )
    warnings = mapping_warnings + warnings

    for warning in warnings:
        print(f"WARNING: {warning}")
    if errors:
        for error in errors:
            print(f"ERROR: {error}")
        raise SystemExit(1)

    if args.dry_run:
        print(f"Parsed {len(parsed.item_rows)} item rows.")
        print(f"Parsed {len(parsed.bidder_bids)} bidder bids.")
        print(f"Parsed {len(parsed.bidder_items)} bidder item rows.")
        print(f"Parsed {len(match_candidates)} match candidate rows.")
        return

    write_csv(args.staging_items, parsed.item_rows, STAGING_ITEM_FIELDS)
    write_csv(args.staging_bidder_bids, parsed.bidder_bids, BIDDER_BID_FIELDS)
    write_csv(args.staging_bidder_items, parsed.bidder_items, BIDDER_ITEM_FIELDS)
    write_csv(args.staging_match_candidates, match_candidates, MATCH_CANDIDATE_FIELDS)
    if getattr(args, "staging_reconciliation", None):
        write_csv(args.staging_reconciliation, [{"warning": warning} for warning in warnings], ["warning"])

    existing_sources = [row for row in normalize_existing_rows(read_csv(args.sources), SOURCE_FIELDS) if row["source_id"] != args.source_id]
    existing_projects = [row for row in normalize_existing_rows(read_csv(args.projects), PROJECT_FIELDS) if row["source_id"] != args.source_id]
    existing_observations = [row for row in normalize_existing_rows(read_csv(args.observations), OBSERVATION_FIELDS) if row["source_id"] != args.source_id]
    existing_bids = [row for row in normalize_existing_rows(read_csv(args.bidder_bids), BIDDER_BID_FIELDS) if row["source_id"] != args.source_id]
    existing_bidder_items = [row for row in normalize_existing_rows(read_csv(args.bidder_items), BIDDER_ITEM_FIELDS) if row["source_id"] != args.source_id]
    existing_bid_tab_items = [row for row in normalize_existing_rows(read_csv(args.bid_tab_items), BID_TAB_ITEM_FIELDS) if row["source_id"] != args.source_id]

    write_csv(args.sources, existing_sources + [source_row(args)], SOURCE_FIELDS)
    write_csv(args.projects, existing_projects + [project_row(args, parsed)], PROJECT_FIELDS)
    write_csv(args.observations, existing_observations + ([] if args.source_only else parsed.observations), OBSERVATION_FIELDS)
    write_csv(args.bidder_bids, existing_bids + parsed.bidder_bids, BIDDER_BID_FIELDS)
    write_csv(args.bidder_items, existing_bidder_items + parsed.bidder_items, BIDDER_ITEM_FIELDS)
    write_csv(args.bid_tab_items, existing_bid_tab_items + parsed.item_rows, BID_TAB_ITEM_FIELDS)

    print(f"Wrote {len(parsed.item_rows)} staging item rows.")
    print(f"Wrote {0 if args.source_only else len(parsed.observations)} promoted item observations.")
    print(f"Wrote {len(parsed.bidder_bids)} bidder bids.")
    print(f"Wrote {len(parsed.bidder_items)} bidder item observations.")
    print(f"Wrote {len(parsed.item_rows)} bid tab item rows.")
    print(f"Wrote {len(match_candidates)} match candidate rows.")


def promote_master_config(args: argparse.Namespace) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise SystemExit("Missing dependency: openpyxl. Run with the bundled Codex Python runtime.") from error

    config_path = Path(args.master_config)
    config_data = json.loads(config_path.read_text(encoding="utf-8"))
    workbook_path = Path(args.workbook)
    workbook = load_workbook(workbook_path, data_only=True, read_only=False)
    selected_sheet = args.sheet
    total_items = total_bids = total_bidder_items = 0
    for source in selected_master_sources(config_data, selected_sheet):
        if source["sheet"] not in workbook.sheetnames:
            raise ValueError(f"Configured sheet {source['sheet']!r} is missing from {workbook_path.name}.")
        parsed = parse_master_sheet(workbook_path, workbook[source["sheet"]], source)
        source_args = argparse.Namespace(**vars(args))
        for key, value in source.items():
            setattr(source_args, key.replace("source_date", "date_basis"), value)
        source_args.workbook = workbook_path
        source_args.source_year = source["source_date"][:4]
        source_args.state = "CO"
        source_args.staging_items = Path("public/data/imports") / f"{source['source_id']}_item_unit_costs.csv"
        source_args.staging_bidder_bids = Path("public/data/imports") / f"{source['source_id']}_bidder_bids.csv"
        source_args.staging_bidder_items = Path("public/data/imports") / f"{source['source_id']}_bidder_item_observations.csv"
        source_args.staging_match_candidates = Path("public/data/imports") / f"{source['source_id']}_match_candidates.csv"
        source_args.staging_reconciliation = Path("public/data/imports") / f"{source['source_id']}_reconciliation.csv"
        promote(source_args, parsed)
        total_items += len(parsed.item_rows)
        total_bids += len(parsed.bidder_bids)
        total_bidder_items += len(parsed.bidder_items)
    print(
        f"Master workbook totals: {total_items} source items, {total_bids} bids, "
        f"{total_bidder_items} bidder item prices."
    )


def selected_master_sources(config_data: dict[str, Any], sheet_name: str | None) -> list[dict[str, Any]]:
    sources = config_data["sources"]
    selected = [source for source in sources if not sheet_name or source["sheet"] == sheet_name]
    if sheet_name and not selected:
        raise ValueError(f"Sheet {sheet_name!r} is not configured for batch import.")
    return selected


def main() -> None:
    parser = argparse.ArgumentParser(description="Import SAQ-style public bid tab workbooks into app CSV data.")
    parser.add_argument("--workbook", default=DEFAULT_WORKBOOK, type=Path)
    parser.add_argument("--master-config", type=Path)
    parser.add_argument("--sheet", help="Import only one configured master-workbook sheet.")
    parser.add_argument("--source-id", default=DEFAULT_SOURCE_ID)
    parser.add_argument("--source-label", default=DEFAULT_SOURCE_LABEL)
    parser.add_argument("--source-type", default=DEFAULT_SOURCE_TYPE)
    parser.add_argument("--source-year", default=DEFAULT_SOURCE_YEAR)
    parser.add_argument("--project-id", default=DEFAULT_PROJECT_ID)
    parser.add_argument("--row-prefix", default=DEFAULT_ROW_PREFIX)
    parser.add_argument("--date-basis", default=DEFAULT_DATE_BASIS)
    parser.add_argument("--agency-owner", default="Town of Breckenridge")
    parser.add_argument("--state", default="CO")
    parser.add_argument("--county-region", default="Summit County")
    parser.add_argument("--work-type", default="Roadway")
    parser.add_argument("--district", default="")
    parser.add_argument("--sources", default=DEFAULT_SOURCES, type=Path)
    parser.add_argument("--projects", default=DEFAULT_PROJECTS, type=Path)
    parser.add_argument("--observations", default=DEFAULT_OBSERVATIONS, type=Path)
    parser.add_argument("--bidder-bids", default=DEFAULT_BIDDER_BIDS, type=Path)
    parser.add_argument("--bidder-items", default=DEFAULT_BIDDER_ITEMS, type=Path)
    parser.add_argument("--bid-tab-items", default=DEFAULT_BID_TAB_ITEMS, type=Path)
    parser.add_argument("--agency-items", default=DEFAULT_AGENCY_ITEMS, type=Path)
    parser.add_argument("--staging-items", default=DEFAULT_STAGING_ITEMS, type=Path)
    parser.add_argument("--staging-bidder-bids", default=DEFAULT_STAGING_BIDDER_BIDS, type=Path)
    parser.add_argument("--staging-bidder-items", default=DEFAULT_STAGING_BIDDER_ITEMS, type=Path)
    parser.add_argument("--staging-match-candidates", default=DEFAULT_STAGING_MATCH_CANDIDATES, type=Path)
    parser.add_argument("--allow-missing-item-codes", action="store_true")
    parser.add_argument("--source-only", action="store_true")
    parser.add_argument("--dry-run", action="store_true")
    args = parser.parse_args()

    if args.master_config:
        promote_master_config(args)
    else:
        promote(args)


if __name__ == "__main__":
    main()
